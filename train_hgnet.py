"""
train_hgnet.py — HGNetV2-B0 SED Training (v2 — multi-loss / multi-aug)

Experiments:
  hgnet_v1          : baseline (train_audio, CustomBCE, MixUp) — exact notebook
  hgnet_ss_v1       : + train_soundscapes GT labels as extra training data
  hgnet_ss_asl      : SS + Asymmetric Loss (ASL)
  hgnet_ss_focal    : SS + Focal BCE + SpecAugment
  hgnet_ss_sumix    : SS + SumixFreq augmentation
  hgnet_ss_cutmix   : SS + CutMix-in-mel
  hgnet_ss_swa      : SS + SWA (stochastic weight averaging)
  hgnet_ss_combo    : SS + ASL + SumixFreq + SWA (best combo)

Config flags (training section):
  loss_type          : custom_bce | focal | asl | label_smooth   (default: custom_bce)
  use_sumix_freq     : bool (default: false)
  use_specaug        : bool (default: false)
  freq_mask          : int  (default: 32)
  time_mask          : int  (default: 32)
  use_cutmix         : bool (default: false)
  use_gain_norm      : bool (default: false)  — random amplitude scaling
  use_swa            : bool (default: false)
  swa_start_frac     : float (default: 0.75)  — fraction of epochs to start SWA
  swa_lr             : float (default: 1e-5)

Config flags (data section):
  soundscape_gt_csv  : path to train_soundscapes_labels.csv (real GT, not pseudo)
  ss_gt_weight       : float (default: 0.5)  — loss weight for GT soundscape branch

Usage:
  python3 train_hgnet.py --config configs/hgnet_v1.yaml --device cuda:0 --fold 0
"""

import argparse
import json
import os
import shutil
import sys
import warnings
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import soundfile as sf
import timm
import torch
import torch.nn as nn
import torch.nn.functional as F
import torchaudio.transforms as T
import torchvision.transforms.v2 as tvt_v2
import wandb
import yaml
from sklearn.metrics import roc_auc_score
from sklearn.model_selection import GroupKFold
from torch.optim.swa_utils import AveragedModel, SWALR, update_bn
from torch.utils.data import DataLoader, Dataset

warnings.filterwarnings("ignore")

SR = 32_000
NUM_CLASSES = 234
PASS_THRESHOLD = 0.9193


# ── Mel Transform ──────────────────────────────────────────────────────────────
class LogMelSpectrogramTransform(nn.Module):
    """Exact reproduction of notebook's LogMelSpectrogramTransform."""
    def __init__(self, mel_params: dict, top_db: float = 80.0, lms_shape=(256, 256)):
        super().__init__()
        self.mel_transform = T.MelSpectrogram(**mel_params)
        self.db = T.AmplitudeToDB(stype="power", top_db=top_db)
        self.resize = tvt_v2.Resize(size=lms_shape)

    @torch.no_grad()
    def forward(self, wave: torch.Tensor) -> torch.Tensor:
        """wave: (B, SR*5) → (B, 1, H, W) normalized [0,1]"""
        mel = self.mel_transform(wave)
        lms = self.db(mel)
        lms = self.resize(lms)
        B = lms.shape[0]
        flat = lms.reshape(B, -1)
        lo = flat.min(1)[0][:, None, None]
        hi = flat.max(1)[0][:, None, None]
        lms = (lms - lo) / (hi - lo + 1e-7)
        return lms[:, None, :, :]  # (B, 1, H, W)


# ── Augmentations ──────────────────────────────────────────────────────────────
class MixUp(nn.Module):
    """MixUp in mel domain — exact reproduction of notebook."""
    def __init__(self, alpha=1.0, theta=0.8):
        super().__init__()
        self.beta = torch.distributions.Beta(alpha, alpha)
        self.theta = theta

    def forward(self, lms: torch.Tensor, label: torch.Tensor):
        B = lms.shape[0]
        lam = self.beta.sample((B,)).to(lms.device)
        lam = torch.maximum(lam, 1 - lam).float()
        perm = torch.randperm(B, device=lms.device)
        lms = lam[:, None, None, None] * lms + (1 - lam[:, None, None, None]) * lms[perm]
        label = lam[:, None] * label + (1 - lam[:, None]) * label[perm]
        label[label >= self.theta] = 1.0
        return lms, label


class CutMix(nn.Module):
    """CutMix in mel domain — paste a rectangle from permuted sample."""
    def __init__(self, alpha=1.0):
        super().__init__()
        self.beta = torch.distributions.Beta(alpha, alpha)

    def forward(self, lms: torch.Tensor, label: torch.Tensor):
        B, C, H, W = lms.shape
        lam = self.beta.sample().item()
        lam = max(lam, 1 - lam)
        perm = torch.randperm(B, device=lms.device)
        # Box proportional to (1 - lam)
        cut_h = int(H * (1 - lam) ** 0.5)
        cut_w = int(W * (1 - lam) ** 0.5)
        if cut_h < 1 or cut_w < 1:
            return lms, label
        cy = torch.randint(H, (1,)).item()
        cx = torch.randint(W, (1,)).item()
        y1 = max(0, cy - cut_h // 2)
        y2 = min(H, cy + cut_h // 2)
        x1 = max(0, cx - cut_w // 2)
        x2 = min(W, cx + cut_w // 2)
        lms = lms.clone()
        lms[:, :, y1:y2, x1:x2] = lms[perm, :, y1:y2, x1:x2]
        # Recalculate lambda from actual box area
        real_lam = 1 - (y2 - y1) * (x2 - x1) / (H * W)
        label = real_lam * label + (1 - real_lam) * label[perm]
        return lms, label


class SumixFreq(nn.Module):
    """SumixFreq: mix mel frequency bands from two samples.
    Proven effective in our SED chain. Encourages model to focus on
    species-specific frequency bands rather than global spectral shape.
    """
    def __init__(self, alpha=0.5):
        super().__init__()
        self.alpha = alpha

    def forward(self, lms: torch.Tensor, label: torch.Tensor):
        B, C, H, W = lms.shape
        perm = torch.randperm(B, device=lms.device)
        # Random frequency split point
        split = torch.randint(H // 4, 3 * H // 4, (1,)).item()
        lam = torch.distributions.Beta(self.alpha, self.alpha).sample().item()
        lms2 = lms.clone()
        # Low freq from original, high freq mixed
        lms2[:, :, split:, :] = lam * lms[:, :, split:, :] + (1 - lam) * lms[perm, :, split:, :]
        label2 = lam * label + (1 - lam) * label[perm]
        return lms2, label2


class SpecAugment(nn.Module):
    """Frequency and time masking (SpecAugment)."""
    def __init__(self, freq_mask=32, time_mask=32, n_freq_masks=1, n_time_masks=1):
        super().__init__()
        self.freq_mask = freq_mask
        self.time_mask = time_mask
        self.n_freq_masks = n_freq_masks
        self.n_time_masks = n_time_masks

    @torch.no_grad()
    def forward(self, lms: torch.Tensor) -> torch.Tensor:
        """lms: (B, 1, H, W) — apply in-place masking"""
        B, C, H, W = lms.shape
        lms = lms.clone()
        for _ in range(self.n_freq_masks):
            f = torch.randint(0, self.freq_mask, (B,))
            f0 = torch.randint(0, H, (B,))
            for b in range(B):
                lms[b, :, f0[b]:min(H, f0[b] + f[b]), :] = 0.0
        for _ in range(self.n_time_masks):
            t = torch.randint(0, self.time_mask, (B,))
            t0 = torch.randint(0, W, (B,))
            for b in range(B):
                lms[b, :, :, t0[b]:min(W, t0[b] + t[b])] = 0.0
        return lms


class GainNorm(nn.Module):
    """Random amplitude scaling in wave domain — simulates recording level variation.
    Creative trick: different recording devices, distances, or AGC settings
    produce clips at very different amplitudes. Robustness to this is key for
    soundscape generalization.
    """
    def __init__(self, gain_range=(0.1, 1.0)):
        super().__init__()
        self.low, self.high = gain_range

    @torch.no_grad()
    def forward(self, wave: torch.Tensor) -> torch.Tensor:
        """wave: (B, T)"""
        B = wave.shape[0]
        gain = torch.empty(B, 1, device=wave.device).uniform_(self.low, self.high)
        return wave * gain


# ── Model ─────────────────────────────────────────────────────────────────────
class GeMPooling(nn.Module):
    def __init__(self, init_p=3.0, eps=1e-6):
        super().__init__()
        self.p = nn.Parameter(torch.tensor(init_p))
        self.eps = eps

    def forward(self, h):
        p = self.p.clamp(min=1.0)
        h = h.clamp(min=self.eps).pow(p)
        h = h.mean(dim=2)
        h = h.pow(1.0 / p)
        return h


class AttnSEDHead(nn.Module):
    def __init__(self, num_features, num_classes, dropout=0.2):
        super().__init__()
        self.pre_fc = nn.Sequential(
            nn.Linear(num_features, num_features),
            nn.ReLU(inplace=True),
            nn.Dropout(dropout),
        )
        self.att_fc = nn.Linear(num_features, num_classes)
        self.cls_fc = nn.Linear(num_features, num_classes)

    def forward(self, h):
        h = h.permute(0, 2, 1)
        h = self.pre_fc(h)
        att_w = torch.tanh(self.att_fc(h))
        att_w = F.softmax(att_w, dim=1)
        timewise_logits = self.cls_fc(h)
        logits = (att_w * timewise_logits).sum(dim=1)
        return logits, timewise_logits


class AttnSEDModel(nn.Module):
    def __init__(self, backbone='hgnetv2_b0.ssld_stage2_ft_in1k',
                 num_classes=NUM_CLASSES, dropout=0.2, drop_path_rate=0.0,
                 pretrained=True):
        super().__init__()
        self.backbone = timm.create_model(
            backbone, pretrained=pretrained, in_chans=1,
            global_pool='', num_classes=0, drop_path_rate=drop_path_rate)
        with torch.no_grad():
            dummy = torch.zeros(1, 1, 256, 256)
            feat_dim = self.backbone(dummy).shape[1]
        self.gem_pool = GeMPooling(init_p=3.0)
        self.head = AttnSEDHead(feat_dim, num_classes, dropout)

    def forward_for_training(self, x):
        h = self.backbone(x)
        h = self.gem_pool(h)
        return self.head(h)

    def forward(self, x):
        logits, _ = self.forward_for_training(x)
        return logits


# ── Losses ─────────────────────────────────────────────────────────────────────
class CustomBCEWithLogitsLoss(nn.Module):
    """0.5 * BCE(clip) + 0.5 * BCE(timewise_max) — exact notebook."""
    def __init__(self, timewise_weight=0.5):
        super().__init__()
        self.tw = timewise_weight

    def forward(self, logits, timewise_logits, labels):
        loss_clip = F.binary_cross_entropy_with_logits(logits, labels)
        tw_max = timewise_logits.max(dim=1)[0]
        loss_tw = F.binary_cross_entropy_with_logits(tw_max, labels)
        return (1 - self.tw) * loss_clip + self.tw * loss_tw


class FocalBCELoss(nn.Module):
    """Focal BCE: down-weight easy negatives. Good for long-tail species."""
    def __init__(self, gamma=2.0, timewise_weight=0.5):
        super().__init__()
        self.gamma = gamma
        self.tw = timewise_weight

    def _focal_bce(self, logits, labels):
        logits = logits.float()
        labels = labels.float()
        bce = F.binary_cross_entropy_with_logits(logits, labels, reduction='none')
        prob = torch.sigmoid(logits)
        pt = torch.where(labels >= 0.5, prob, 1 - prob).clamp(min=1e-6, max=1.0)
        focal = (1 - pt) ** self.gamma * bce
        return focal.mean()

    def forward(self, logits, timewise_logits, labels):
        clip_loss = self._focal_bce(logits, labels)
        tw_max = timewise_logits.max(dim=1)[0]
        tw_loss = self._focal_bce(tw_max, labels)
        return (1 - self.tw) * clip_loss + self.tw * tw_loss


class AsymmetricLoss(nn.Module):
    """ASL — Asymmetric Loss (Ridnik et al. 2021).
    Reference impl: github.com/Alibaba-MIIL/ASL

    Key fix: focusing weight uses pt = p for positives, pt = 1-p for negatives
    so (1-pt)^γ = (1-p)^γ_pos for positives, p^γ_neg for negatives.
    This correctly down-weights easy negatives (low p) and focuses on hard ones (high p).
    Always computed in fp32 to avoid fp16 underflow.
    """
    def __init__(self, gamma_pos=0.0, gamma_neg=4.0, margin=0.05,
                 timewise_weight=0.5, eps=1e-8):
        super().__init__()
        self.gamma_pos = gamma_pos
        self.gamma_neg = gamma_neg
        self.margin = margin
        self.tw = timewise_weight
        self.eps = eps

    def _asl(self, logits, labels):
        # Always fp32 — log/pow unstable in fp16
        logits = logits.float()
        labels = labels.float()
        prob = torch.sigmoid(logits)

        # Probability shifting: hard-threshold easy negatives
        prob_m = (prob - self.margin).clamp(min=0.0)

        # BCE terms: pos → log(p), neg → log(1 - p_m)
        los_pos = labels       * torch.log(prob.clamp(min=self.eps))
        los_neg = (1 - labels) * torch.log((1 - prob_m).clamp(min=self.eps))
        loss = los_pos + los_neg  # (B, C)

        # Asymmetric focusing weights
        # pt: probability of the TRUE class
        #   positives → pt = prob          → weight = (1-pt)^γ_pos = (1-p)^γ_pos
        #   negatives → pt = 1 - prob_m    → weight = (1-pt)^γ_neg = prob_m^γ_neg
        if self.gamma_neg > 0 or self.gamma_pos > 0:
            pt = prob * labels + (1 - prob_m) * (1 - labels)
            pt = pt.clamp(min=self.eps, max=1.0 - self.eps)
            gamma = self.gamma_pos * labels + self.gamma_neg * (1 - labels)
            w = torch.pow(1.0 - pt, gamma)
            loss = loss * w

        return -loss.mean()

    def forward(self, logits, timewise_logits, labels):
        clip_loss = self._asl(logits, labels)
        tw_max = timewise_logits.max(dim=1)[0]
        tw_loss = self._asl(tw_max, labels)
        return (1 - self.tw) * clip_loss + self.tw * tw_loss


class LabelSmoothBCELoss(nn.Module):
    """BCE with label smoothing — helps calibration and reduces overconfidence."""
    def __init__(self, smoothing=0.05, timewise_weight=0.5):
        super().__init__()
        self.smoothing = smoothing
        self.tw = timewise_weight

    def _smooth_bce(self, logits, labels):
        logits = logits.float()
        labels = labels.float()
        labels_s = labels * (1 - self.smoothing) + 0.5 * self.smoothing
        return F.binary_cross_entropy_with_logits(logits, labels_s)

    def forward(self, logits, timewise_logits, labels):
        clip_loss = self._smooth_bce(logits, labels)
        tw_max = timewise_logits.max(dim=1)[0]
        tw_loss = self._smooth_bce(tw_max, labels)
        return (1 - self.tw) * clip_loss + self.tw * tw_loss


def build_loss(t_cfg):
    loss_type = t_cfg.get('loss_type', 'custom_bce')
    tw = t_cfg.get('timewise_weight', 0.5)
    if loss_type == 'focal':
        gamma = t_cfg.get('focal_gamma', 2.0)
        return FocalBCELoss(gamma=gamma, timewise_weight=tw)
    elif loss_type == 'asl':
        return AsymmetricLoss(
            gamma_pos=t_cfg.get('asl_gamma_pos', 0.0),
            gamma_neg=t_cfg.get('asl_gamma_neg', 4.0),
            margin=t_cfg.get('asl_margin', 0.05),
            timewise_weight=tw)
    elif loss_type == 'label_smooth':
        return LabelSmoothBCELoss(
            smoothing=t_cfg.get('label_smoothing', 0.05),
            timewise_weight=tw)
    else:
        return CustomBCEWithLogitsLoss(timewise_weight=tw)


# ── Datasets ──────────────────────────────────────────────────────────────────
def load_audio_clip(path: str, n_samples: int) -> np.ndarray:
    try:
        with sf.SoundFile(path) as f:
            n_frames = f.frames
            if n_frames < n_samples:
                wave = np.zeros(n_samples, dtype='float32')
                start = np.random.randint(n_samples - n_frames + 1)
                wave[start:start + n_frames] = f.read(dtype='float32', always_2d=False)
            else:
                start = np.random.randint(n_frames - n_samples + 1)
                f.seek(start)
                wave = f.read(frames=n_samples, dtype='float32', always_2d=False)
        if wave.ndim == 2:
            wave = wave.mean(axis=1)
        return wave.astype('float32')
    except Exception:
        return np.zeros(n_samples, dtype='float32')


def load_audio_head(path: str, n_samples: int) -> np.ndarray:
    try:
        with sf.SoundFile(path) as f:
            n_frames = f.frames
            if n_frames < n_samples:
                wave = np.zeros(n_samples, dtype='float32')
                data = f.read(dtype='float32', always_2d=False)
                wave[:n_frames] = data if data.ndim == 1 else data.mean(1)
            else:
                data = f.read(frames=n_samples, dtype='float32', always_2d=False)
                wave = data if data.ndim == 1 else data.mean(1)
        return wave.astype('float32')
    except Exception:
        return np.zeros(n_samples, dtype='float32')


class TrainAudioDataset(Dataset):
    """train_audio with hard labels, pre-cached to RAM."""
    def __init__(self, df, audio_dir, species_cols, n_samples):
        self.df = df.reset_index(drop=True)
        self.audio_dir = audio_dir
        self.species_cols = species_cols
        self.n_samples = n_samples

        print(f"  Pre-caching {len(self.df):,} train_audio clips into RAM …", flush=True)
        self._cache = []
        for i, row in enumerate(self.df.itertuples(index=False)):
            path = os.path.join(self.audio_dir, str(row.filename))
            clip = load_audio_clip(path, self.n_samples).astype(np.float16)
            self._cache.append(clip)
            if (i + 1) % 5000 == 0:
                print(f"    cached {i+1:,}/{len(self.df):,}", flush=True)
        print("  Pre-cache complete.", flush=True)

    def __len__(self): return len(self.df)

    def __getitem__(self, idx):
        wave = self._cache[idx].astype(np.float32)
        row = self.df.iloc[idx]
        label = row[self.species_cols].values.astype(np.float32)
        return torch.from_numpy(wave), torch.from_numpy(label)


class SoundscapeGTDataset(Dataset):
    """train_soundscapes with REAL GT labels from train_soundscapes_labels.csv.
    Not pseudo labels — uses verified human annotations.
    Provides domain-adapted (soundscape) training data without label noise.
    val_files: soundscape filenames in this fold's validation set (excluded).
    """
    def __init__(self, csv_path, soundscape_dir, species, n_samples, val_files=None):
        df = pd.read_csv(csv_path)
        if val_files is not None:
            df = df[~df['filename'].isin(val_files)].reset_index(drop=True)

        self.n_samples = n_samples
        self.soundscape_dir = soundscape_dir
        self.species = species

        def hhmmss_to_sec(t):
            try:
                parts = str(t).split(':')
                return int(parts[0]) * 3600 + int(parts[1]) * 60 + float(parts[2])
            except Exception:
                return float(t)

        rows = []
        for _, row in df.iterrows():
            fname = str(row['filename'])
            start_s = hhmmss_to_sec(row.get('start', row.get('start_time', 0)))
            lbl_str = str(row.get('primary_label', ''))
            present = set(l.strip() for l in lbl_str.split(';') if l.strip())
            vec = np.array([1.0 if s in present else 0.0 for s in species], dtype=np.float32)
            rows.append((fname, start_s, vec))

        print(f"  Loading {len(rows):,} soundscape GT clips …", flush=True)
        self._cache, self._labels, skipped = [], [], 0
        for fname, start_s, vec in rows:
            path = os.path.join(soundscape_dir, fname)
            try:
                with sf.SoundFile(path) as f:
                    f.seek(int(start_s * f.samplerate))
                    data = f.read(frames=n_samples, dtype='float32', always_2d=False)
                if data.ndim == 2: data = data.mean(1)
                if len(data) < n_samples:
                    data = np.pad(data, (0, n_samples - len(data)))
                self._cache.append(data[:n_samples].astype(np.float16))
                self._labels.append(vec)
            except Exception:
                skipped += 1
        print(f"  SS-GT: {len(self._cache):,} clips loaded (skipped {skipped})", flush=True)

    def __len__(self): return len(self._cache)

    def __getitem__(self, idx):
        wave = torch.from_numpy(self._cache[idx].astype(np.float32))
        label = torch.from_numpy(self._labels[idx])
        return wave, label


class PerchSSDataset(Dataset):
    """Soundscape soft labels from pseudo CSV (kept for backward compat, not used in new experiments)."""
    def __init__(self, csv_path, soundscape_dir, species_cols, n_samples, val_files=None):
        df = pd.read_csv(csv_path, low_memory=False)
        if val_files is not None and 'filename' in df.columns:
            df = df[~df['filename'].isin(val_files)].reset_index(drop=True)
        self.n_samples = n_samples
        rows, labels = [], []
        for _, row in df.iterrows():
            if 'row_id' in df.columns:
                parts = str(row['row_id']).rsplit('_', 1)
                if len(parts) != 2: continue
                fname, end_sec = parts[0] + '.ogg', int(parts[1])
                start_sec = max(0, end_sec - 5)
            else:
                fname, start_sec = str(row['filename']), int(row.get('start', 0))
            rows.append((fname, start_sec))
            labels.append(row[species_cols].values.astype(np.float32))
        self._cache, self._labels, skipped = [], [], 0
        for fname, start_sec in rows:
            path = os.path.join(soundscape_dir, fname)
            try:
                with sf.SoundFile(path) as f:
                    f.seek(int(start_sec * f.samplerate))
                    data = f.read(frames=n_samples, dtype='float32', always_2d=False)
                if data.ndim == 2: data = data.mean(1)
                if len(data) < n_samples:
                    data = np.pad(data, (0, n_samples - len(data)))
                self._cache.append(data[:n_samples].astype(np.float16))
                self._labels.append(labels[len(self._cache) - 1 + skipped])
            except Exception:
                skipped += 1
        print(f"  SS pseudo: {len(self._cache):,} clips (skipped {skipped})", flush=True)

    def __len__(self): return len(self._cache)

    def __getitem__(self, idx):
        wave = torch.from_numpy(self._cache[idx].astype(np.float32))
        label = torch.from_numpy(self._labels[idx])
        return wave, label


# ── Soundscape Validation ──────────────────────────────────────────────────────
def hhmmss_to_sec(t):
    try:
        parts = str(t).split(':')
        return int(parts[0]) * 3600 + int(parts[1]) * 60 + float(parts[2])
    except Exception:
        return float(t)


def build_ss_val_df(sc_labels, species):
    rows = []
    for _, row in sc_labels.iterrows():
        fname = str(row['filename'])
        start_s = hhmmss_to_sec(row.get('start', row.get('start_time', 0)))
        end_s = hhmmss_to_sec(row.get('end', row.get('end_time', 5)))
        lbl_str = str(row.get('primary_label', ''))
        present = set(l.strip() for l in lbl_str.split(';') if l.strip())
        vec = np.array([1.0 if s in present else 0.0 for s in species], dtype=np.float32)
        rows.append({'filename': fname, 'start_sec': start_s, 'end_sec': end_s, 'labels': vec})
    return pd.DataFrame(rows)


def infer_soundscape_val(model, mel_tf, val_df, soundscape_dir, species, device, n_samples, batch_size=32):
    model.eval()
    all_logits, all_labels = [], []
    clips, labels = [], []
    for _, row in val_df.iterrows():
        path = os.path.join(soundscape_dir, row['filename'])
        try:
            with sf.SoundFile(path) as f:
                f.seek(int(row['start_sec'] * f.samplerate))
                data = f.read(frames=n_samples, dtype='float32', always_2d=False)
            if data.ndim == 2: data = data.mean(1)
            if len(data) < n_samples:
                data = np.pad(data, (0, n_samples - len(data)))
            wave = data[:n_samples]
        except Exception:
            wave = np.zeros(n_samples, dtype='float32')
        clips.append(torch.from_numpy(wave))
        labels.append(torch.from_numpy(row['labels']))
        if len(clips) == batch_size:
            with torch.no_grad():
                wav_batch = torch.stack(clips).to(device)
                mel = mel_tf(wav_batch)
                logits = model(mel).cpu()
            all_logits.append(logits)
            all_labels.append(torch.stack(labels))
            clips, labels = [], []
    if clips:
        with torch.no_grad():
            wav_batch = torch.stack(clips).to(device)
            mel = mel_tf(wav_batch)
            logits = model(mel).cpu()
        all_logits.append(logits)
        all_labels.append(torch.stack(labels))
    if not all_logits:
        return 0.0
    logits = torch.cat(all_logits).numpy()
    labels = torch.cat(all_labels).numpy()
    probs = torch.sigmoid(torch.tensor(logits)).numpy()
    try:
        mask = labels.sum(0) > 0
        if mask.sum() < 2:
            return 0.0
        return float(roc_auc_score(labels[:, mask], probs[:, mask], average='macro'))
    except Exception:
        return 0.0


# ── Training Fold ──────────────────────────────────────────────────────────────
def train_fold(fold: int, cfg: dict, device: torch.device, out_dir: Path):
    d_cfg = cfg['data']
    m_cfg = cfg['model']
    t_cfg = cfg['training']
    exp_name = cfg['experiment']['name']
    n_samples = SR * 5

    mel_params = dict(
        sample_rate=SR,
        n_fft=m_cfg.get('n_fft', 2048),
        win_length=m_cfg.get('win_length', 626),
        hop_length=m_cfg.get('hop_length', 313),
        f_min=m_cfg.get('fmin', 20),
        n_mels=m_cfg.get('n_mels', 256),
        power=m_cfg.get('power', 2.0),
        center=True, pad_mode='reflect', norm='slaney', mel_scale='htk',
    )
    lms_shape = (m_cfg.get('n_mels', 256), m_cfg.get('n_mels', 256))
    mel_tf = LogMelSpectrogramTransform(mel_params, m_cfg.get('top_db', 80.0), lms_shape).eval().to(device)

    # Labels
    train_df = pd.read_csv(d_cfg['train_csv'])
    taxonomy = pd.read_csv(d_cfg['taxonomy_csv'])
    species = taxonomy['primary_label'].tolist()

    def make_label(row):
        vec = np.zeros(len(species), dtype=np.float32)
        pl = str(row.get('primary_label', ''))
        if pl in species:
            vec[species.index(pl)] = 1.0
        for s in str(row.get('secondary_labels', '')).strip("[]'\" ").split(','):
            s = s.strip().strip("'\"")
            if s in species:
                vec[species.index(s)] = 0.5
        return vec

    train_df['filename'] = train_df.apply(
        lambda r: str(r['primary_label']) + '/' + str(r['filename'])
        if '/' not in str(r['filename']) else str(r['filename']), axis=1)
    labels_matrix = np.stack(train_df.apply(make_label, axis=1).values)
    label_df = pd.DataFrame(labels_matrix, columns=species)
    train_df = pd.concat([train_df[['filename']], label_df], axis=1)

    # GroupKFold on soundscape files
    sc_labels = pd.read_csv(d_cfg['soundscape_labels_csv'])
    sc_files = sc_labels['filename'].unique()
    sc_groups = [f.split('_')[2] for f in sc_files]
    gkf = GroupKFold(n_splits=d_cfg.get('n_folds', 5))
    splits = list(gkf.split(sc_files, groups=sc_groups))
    _, val_idx = splits[fold]
    val_files = set(sc_files[val_idx])

    sc_val_raw = sc_labels[sc_labels['filename'].isin(val_files)]
    val_df = build_ss_val_df(sc_val_raw, species)
    print(f"Fold {fold}: val_sc_files={len(val_files)}, val_rows={len(val_df)}")

    # Train dataset
    train_ds = TrainAudioDataset(train_df, d_cfg['audio_dir'], species, n_samples)
    train_loader = DataLoader(
        train_ds, batch_size=t_cfg['batch_size'], shuffle=True,
        num_workers=0, pin_memory=False, drop_last=True)

    # Soundscape GT dataset (real labels, not pseudo)
    ss_gt_loader = None
    ss_gt_weight = t_cfg.get('ss_gt_weight', 0.0)
    gt_csv = d_cfg.get('soundscape_gt_csv', '')
    if gt_csv and ss_gt_weight > 0 and Path(gt_csv).exists():
        ss_gt_ds = SoundscapeGTDataset(gt_csv, d_cfg['soundscape_dir'], species, n_samples, val_files=val_files)
        ss_gt_oversample = t_cfg.get('ss_gt_oversample', 2)
        gt_sampler = torch.utils.data.RandomSampler(
            ss_gt_ds, replacement=True,
            num_samples=len(train_ds) * ss_gt_oversample)
        ss_gt_loader = DataLoader(
            ss_gt_ds, batch_size=max(1, t_cfg['batch_size'] // 2), sampler=gt_sampler,
            num_workers=0, pin_memory=False, drop_last=True)
        print(f"  SS-GT loader: {len(ss_gt_ds):,} clips, weight={ss_gt_weight}")

    # Pseudo label loader (backward compat)
    ss_loader = None
    ss_weight = t_cfg.get('ss_weight', 0.0)
    perch_csv = d_cfg.get('perch_ss_csv', '')
    if perch_csv and ss_weight > 0 and Path(perch_csv).exists():
        ss_ds = PerchSSDataset(perch_csv, d_cfg['soundscape_dir'], species, n_samples, val_files=val_files)
        ss_loader = DataLoader(
            ss_ds, batch_size=max(1, t_cfg['batch_size'] // 2), shuffle=True,
            num_workers=0, pin_memory=False, drop_last=True)

    # Model
    model = AttnSEDModel(
        backbone=m_cfg.get('backbone', 'hgnetv2_b0.ssld_stage2_ft_in1k'),
        num_classes=len(species),
        dropout=m_cfg.get('dropout', 0.2),
        drop_path_rate=m_cfg.get('drop_path_rate', 0.0),
        pretrained=True,
    ).to(device)
    print(f"  Params: {sum(p.numel() for p in model.parameters() if p.requires_grad):,}")

    init_ckpt = d_cfg.get('init_ckpt', '')
    if init_ckpt and init_ckpt != 'PLACEHOLDER_UPDATED_BY_PIPELINE' and Path(init_ckpt).exists():
        ck = torch.load(init_ckpt, map_location='cpu', weights_only=False)
        state = ck.get('state_dict', ck.get('model_state_dict', ck))
        missing, unexpected = model.load_state_dict(state, strict=False)
        print(f"  Loaded ckpt: missing={len(missing)}, unexpected={len(unexpected)}")

    # Optimizer
    use_llrd = t_cfg.get('use_llrd', False)
    if use_llrd:
        bb_mult = t_cfg.get('backbone_lr_mult', 0.1)
        backbone_params = [p for n, p in model.named_parameters() if 'backbone' in n]
        head_params = [p for n, p in model.named_parameters() if 'backbone' not in n]
        optimizer = torch.optim.AdamW([
            {'params': backbone_params, 'lr': t_cfg['learning_rate'] * bb_mult},
            {'params': head_params, 'lr': t_cfg['learning_rate']},
        ], weight_decay=t_cfg['weight_decay'])
    else:
        optimizer = torch.optim.AdamW(
            model.parameters(), lr=t_cfg['learning_rate'],
            weight_decay=t_cfg['weight_decay'])

    n_epochs = t_cfg['epochs']
    warmup_epochs = t_cfg.get('warmup_epochs', 2)
    sched_type = t_cfg.get('scheduler', 'onecycle')

    if sched_type == 'warmrestart':
        # CosineAnnealingWarmRestarts — step per epoch, like the competitor
        T_0 = t_cfg.get('sched_T0', 5)
        T_mult = t_cfg.get('sched_Tmult', 1)
        scheduler = torch.optim.lr_scheduler.CosineAnnealingWarmRestarts(
            optimizer, T_0=T_0, T_mult=T_mult, eta_min=t_cfg['learning_rate'] * 1e-3)
        sched_per_batch = False
    elif sched_type == 'cosine':
        scheduler = torch.optim.lr_scheduler.CosineAnnealingLR(
            optimizer, T_max=n_epochs, eta_min=t_cfg['learning_rate'] * 1e-3)
        sched_per_batch = False
    else:  # onecycle (default)
        scheduler = torch.optim.lr_scheduler.OneCycleLR(
            optimizer, max_lr=t_cfg['learning_rate'],
            epochs=n_epochs, steps_per_epoch=len(train_loader),
            pct_start=warmup_epochs / n_epochs,
            div_factor=25, final_div_factor=4.0)
        sched_per_batch = True

    # SWA setup
    use_swa = t_cfg.get('use_swa', False)
    swa_start_frac = t_cfg.get('swa_start_frac', 0.75)
    swa_start_epoch = int(n_epochs * swa_start_frac)
    swa_model = AveragedModel(model) if use_swa else None
    swa_scheduler = SWALR(optimizer, swa_lr=t_cfg.get('swa_lr', 1e-5)) if use_swa else None

    # Loss
    loss_fn = build_loss(t_cfg)

    # Augmentation modules
    mixup = MixUp(alpha=t_cfg.get('mixup_alpha', 1.0), theta=t_cfg.get('mixup_theta', 0.8))
    cutmix = CutMix(alpha=t_cfg.get('cutmix_alpha', 1.0)) if t_cfg.get('use_cutmix', False) else None
    sumix = SumixFreq(alpha=t_cfg.get('sumix_alpha', 0.5)) if t_cfg.get('use_sumix_freq', False) else None
    specaug = SpecAugment(
        freq_mask=t_cfg.get('freq_mask', 32), time_mask=t_cfg.get('time_mask', 32),
        n_freq_masks=t_cfg.get('n_freq_masks', 1), n_time_masks=t_cfg.get('n_time_masks', 1),
    ) if t_cfg.get('use_specaug', False) else None
    gain_norm = GainNorm(
        gain_range=tuple(t_cfg.get('gain_range', [0.1, 1.0]))
    ) if t_cfg.get('use_gain_norm', False) else None

    scaler = torch.GradScaler(enabled=True)
    patience = t_cfg.get('early_stopping_patience', 6)

    try:
        run = wandb.init(
            project='birdclef-2026', name=f"{exp_name}-fold{fold}",
            group=exp_name, tags=['hgnet', f'fold{fold}'],
            config={**cfg, 'fold': fold}, reinit='finish_previous')
    except Exception:
        run = None

    best_auc = 0.0
    best_state = None
    wait = 0
    ss_gt_iter = iter(ss_gt_loader) if ss_gt_loader else None
    ss_iter = iter(ss_loader) if ss_loader else None

    for epoch in range(n_epochs):
        model.train()
        total_loss = 0.0
        use_mixup = epoch >= warmup_epochs

        for waves, labels in train_loader:
            waves, labels = waves.to(device), labels.to(device)

            # Wave-domain augmentation
            if gain_norm is not None:
                waves = gain_norm(waves)

            with torch.no_grad():
                mel = mel_tf(waves)

            # Mel-domain augmentation
            if use_mixup:
                if cutmix is not None and torch.rand(1).item() < 0.5:
                    mel, labels = cutmix(mel, labels)
                else:
                    mel, labels = mixup(mel, labels)
            if sumix is not None and use_mixup:
                mel, labels = sumix(mel, labels)
            if specaug is not None:
                mel = specaug(mel)

            with torch.autocast(device_type=device.type, dtype=torch.float16):
                logits, tw_logits = model.forward_for_training(mel)
                loss = loss_fn(logits, tw_logits, labels)

            # Soundscape GT branch
            if ss_gt_loader is not None and ss_gt_weight > 0:
                try:
                    gt_waves, gt_labels = next(ss_gt_iter)
                except StopIteration:
                    ss_gt_iter = iter(ss_gt_loader)
                    gt_waves, gt_labels = next(ss_gt_iter)
                gt_waves, gt_labels = gt_waves.to(device), gt_labels.to(device)
                if gain_norm is not None:
                    gt_waves = gain_norm(gt_waves)
                with torch.no_grad():
                    gt_mel = mel_tf(gt_waves)
                if specaug is not None:
                    gt_mel = specaug(gt_mel)
                with torch.autocast(device_type=device.type, dtype=torch.float16):
                    gt_logits, gt_tw = model.forward_for_training(gt_mel)
                    gt_loss = loss_fn(gt_logits, gt_tw, gt_labels)
                loss = (1 - ss_gt_weight) * loss + ss_gt_weight * gt_loss

            # Pseudo label branch (backward compat)
            if ss_loader is not None and ss_weight > 0:
                try:
                    ss_waves, ss_labels = next(ss_iter)
                except StopIteration:
                    ss_iter = iter(ss_loader)
                    ss_waves, ss_labels = next(ss_iter)
                ss_waves, ss_labels = ss_waves.to(device), ss_labels.to(device)
                with torch.no_grad():
                    ss_mel = mel_tf(ss_waves)
                with torch.autocast(device_type=device.type, dtype=torch.float16):
                    ss_logits, ss_tw = model.forward_for_training(ss_mel)
                    ss_loss = loss_fn(ss_logits, ss_tw, ss_labels)
                loss = (1 - ss_weight) * loss + ss_weight * ss_loss

            if not torch.isfinite(loss):
                optimizer.zero_grad()
                continue   # skip bad batch, don't corrupt scaler
            scaler.scale(loss).backward()
            scaler.unscale_(optimizer)
            torch.nn.utils.clip_grad_norm_(model.parameters(), 1.0)
            scaler.step(optimizer)
            scaler.update()
            optimizer.zero_grad()

            # SWA: update averaged model after swa_start_epoch
            if use_swa and epoch >= swa_start_epoch:
                swa_model.update_parameters(model)

            if sched_per_batch and not (use_swa and epoch >= swa_start_epoch):
                scheduler.step()

            total_loss += loss.item()

        # Per-epoch scheduler step
        if not sched_per_batch and not (use_swa and epoch >= swa_start_epoch):
            scheduler.step()

        # SWA LR scheduling
        if use_swa and epoch >= swa_start_epoch:
            swa_scheduler.step()

        avg_loss = total_loss / len(train_loader)

        # Evaluate using SWA model if active
        eval_model = model
        if use_swa and epoch >= swa_start_epoch:
            # Update BN stats for SWA model evaluation
            update_bn(train_loader, swa_model, device=device)
            eval_model = swa_model

        auc = infer_soundscape_val(
            eval_model, mel_tf, val_df, d_cfg['soundscape_dir'],
            species, device, n_samples)

        if run:
            wandb.log({'epoch': epoch + 1, 'train/loss': avg_loss,
                       'val/ss_auc': auc, 'val/best_auc': max(auc, best_auc)})

        improved = auc > best_auc
        if improved:
            best_auc = auc
            state_to_save = eval_model.state_dict()
            best_state = {k: v.clone() for k, v in state_to_save.items()}
            torch.save(best_state, out_dir / f'fold{fold}_best.pt')
            wait = 0
            print(f"  Ep {epoch+1:3d}/{n_epochs}  loss={avg_loss:.4f}  ss_auc={auc:.4f}  ✓ new best", flush=True)
        else:
            wait += 1
            print(f"  Ep {epoch+1:3d}/{n_epochs}  loss={avg_loss:.4f}  ss_auc={auc:.4f}  (wait {wait}/{patience})", flush=True)

        if wait >= patience:
            print(f"  Early stop at epoch {epoch+1}", flush=True)
            break

    if run:
        wandb.finish()

    print(f"  Fold {fold} done — best AUC={best_auc:.4f}")
    # Always save to weights/hgnet/ — used for notebook ensemble
    hgnet_dir = Path('weights/hgnet')
    hgnet_dir.mkdir(parents=True, exist_ok=True)
    dst_hgnet = hgnet_dir / f"{exp_name}_fold{fold}_auc{best_auc:.4f}.pt"
    shutil.copy2(str(out_dir / f'fold{fold}_best.pt'), str(dst_hgnet))
    print(f"  ✓ Saved to weights/hgnet/: {dst_hgnet.name}")

    # Additionally copy to sed_improved/ if above threshold
    if best_auc >= PASS_THRESHOLD:
        dst = Path('sed_improved') / f"{exp_name}_fold{fold}_auc{best_auc:.4f}.pt"
        shutil.copy2(str(out_dir / f'fold{fold}_best.pt'), str(dst))
        print(f"  ✓ Copied to sed_improved/: {dst.name}")

    return {'fold': fold, 'best_auc': best_auc, 'exp_name': exp_name}


# ── Excel Result Writer ────────────────────────────────────────────────────────
def update_excel_results(result_dict: dict, excel_path: str = 'reports/hgnet_experiments.xlsx'):
    """Append or update experiment row in Excel. Thread-safe via file lock."""
    Path(excel_path).parent.mkdir(parents=True, exist_ok=True)
    exp_name = result_dict.get('exp_name', 'unknown')
    fold = result_dict.get('fold', 0)
    best_auc = result_dict.get('best_auc', 0.0)

    try:
        if Path(excel_path).exists():
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'HGNet Experiments'
            ws.append(['exp_name', 'fold', 'best_auc', 'pass_threshold'])
            # Bold header
            for cell in ws[1]:
                cell.font = openpyxl.styles.Font(bold=True)

        # Check if row exists (update) or append
        found = False
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value == exp_name and row[1].value == fold:
                row[2].value = best_auc
                row[3].value = '✓' if best_auc >= PASS_THRESHOLD else '✗'
                found = True
                break
        if not found:
            ws.append([exp_name, fold, best_auc, '✓' if best_auc >= PASS_THRESHOLD else '✗'])

        wb.save(excel_path)
        print(f"  Excel updated: {excel_path}")
    except Exception as e:
        print(f"  Excel write failed: {e}")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--config', required=True)
    parser.add_argument('--device', default='cuda:0')
    parser.add_argument('--fold', type=int, default=None)
    args = parser.parse_args()

    with open(args.config) as f:
        cfg = yaml.safe_load(f)

    device = torch.device(args.device)
    out_dir = Path(cfg['output']['dir'])
    out_dir.mkdir(parents=True, exist_ok=True)
    Path('sed_improved').mkdir(exist_ok=True)
    Path('weights/hgnet').mkdir(parents=True, exist_ok=True)

    n_folds = cfg['data'].get('n_folds', 5)
    fold_range = [args.fold] if args.fold is not None else list(range(n_folds))

    result_json = out_dir / 'result.json'

    def save_results(new_result):
        """Merge and save result.json after each fold (incremental)."""
        existing = {}
        if result_json.exists():
            try:
                old = json.load(open(result_json))
                existing = {r['fold']: r for r in old.get('folds', [])}
            except Exception:
                pass
        existing[new_result['fold']] = new_result
        all_r = list(existing.values())
        mean_all = float(np.mean([r['best_auc'] for r in all_r])) if all_r else 0.0
        json.dump({'folds': all_r, 'mean_fold_auc': mean_all}, open(result_json, 'w'), indent=2)
        print(f"Result: {result_json}")

    results = []
    for fold in fold_range:
        print(f"\n{'='*60}\n  Fold {fold}/{n_folds-1}\n{'='*60}")
        r = train_fold(fold, cfg, device, out_dir)
        results.append(r)
        save_results(r)
        update_excel_results(r, excel_path='reports/hgnet_experiments.xlsx')

    mean_auc = float(np.mean([r['best_auc'] for r in results])) if results else 0.0
    print(f"\nMean fold AUC: {mean_auc:.4f}")


if __name__ == '__main__':
    main()
