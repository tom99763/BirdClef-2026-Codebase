"""Experiment monitor for NS chain pipeline.

Watches:
  - Perch head retrain  (outputs/logs/perch_head_retrain.log)
  - SED NS 20s  r1-r4   (outputs/sed-ns-b0-20s-r{N}/, logs/sed_ns_20s_r{N}_fold{F}.log)
  - SSM NS 20s  r1-r4   (outputs/ssm-ns-b0-10s-r{N}/, logs/ssm_ns_10s_r{N}_fold{F}.log)
  - Auto chains         (outputs/logs/auto_{sed,ssm}_ns_20s_full.log)

Usage:
    python scripts/monitor_experiments.py           # print only
    python scripts/monitor_experiments.py --excel   # print + update Excel
"""

import argparse
import json
import os
import re
from datetime import datetime
from pathlib import Path

import pandas as pd

LOG_DIR      = Path("outputs/logs")
OUTPUTS_DIR  = Path("outputs")
EXCEL_PATH   = Path("reports/ns_chain_progress.xlsx")


# ── Parsers ───────────────────────────────────────────────────────────────────

def _last_lines(path, n: int = 200, buf: int = 262144) -> list[str]:
    try:
        with open(path, "rb") as f:
            f.seek(0, 2)
            size = f.tell()
            f.seek(max(0, size - buf))
            return f.read().decode("utf-8", errors="replace").splitlines()[-n:]
    except Exception:
        return []


def parse_perch_log() -> dict:
    lines = _last_lines(LOG_DIR / "perch_head_retrain.log")
    info = {"status": "not_started", "epoch": 0, "best_auc": 0.0, "cur_auc": 0.0, "loss": 0.0}
    pat = re.compile(r"Epoch\s+(\d+)/(\d+)\s*\|\s*loss=([\d.]+)\s*\|\s*val_roc_auc=([\d.]+)")
    best_pat = re.compile(r"best=([\d.]+)")
    for ln in lines:
        m = pat.search(ln)
        if m:
            info["status"]   = "running"
            info["epoch"]    = int(m.group(1))
            info["total_ep"] = int(m.group(2))
            info["loss"]     = float(m.group(3))
            info["cur_auc"]  = float(m.group(4))
        bm = best_pat.search(ln)
        if bm:
            v = float(bm.group(1))
            if v > info["best_auc"]:
                info["best_auc"] = v
    if "Early stopping" in " ".join(lines):
        info["status"] = "early_stopped"
    if any("Training complete" in l or "Saved best model" in l for l in lines[-10:]):
        info["status"] = "done"
    return info


def parse_ns_fold_log(log_path: str) -> dict:
    lines = _last_lines(log_path)
    info = {"epoch": 0, "cur_auc": 0.0, "best_auc": 0.0, "loss": 0.0, "done": False}
    pat = re.compile(r"Ep\s+(\d+)/\s*\d+\s+loss=([\d.]+)\s+ss_auc=([\d.]+)")
    for ln in lines:
        m = pat.search(ln)
        if m:
            info["epoch"]   = int(m.group(1))
            info["loss"]    = float(m.group(2))
            info["cur_auc"] = float(m.group(3))
            if info["cur_auc"] > info["best_auc"]:
                info["best_auc"] = info["cur_auc"]
    if any("Early stopping" in l for l in lines):
        info["done"] = True
    return info


def parse_ns_exp(prefix: str, rounds: int = 4) -> list[dict]:
    """Returns one dict per (round, fold) with status info."""
    rows = []
    for r in range(1, rounds + 1):
        out_dir   = OUTPUTS_DIR / f"{prefix}-ns-b0-20s-r{r}"
        result_f  = out_dir / "result.json"

        # Check if round is fully done
        if result_f.exists():
            try:
                res = json.loads(result_f.read_text())
                mean_auc = res.get("mean_fold_auc", res.get("mean_auc", 0))
                rows.append({
                    "prefix": prefix, "round": r, "fold": "all",
                    "status": "DONE", "epoch": "-",
                    "cur_auc": mean_auc, "best_auc": mean_auc, "loss": "-",
                })
                continue
            except Exception:
                pass

        # Per-fold status
        for fold in range(5):
            ckpt = out_dir / f"fold{fold}_best.pt"
            log_name = f"{prefix}_ns_20s_r{r}_fold{fold}.log"
            log_path = LOG_DIR / log_name

            if ckpt.exists():
                # Fold done — read best AUC from log
                fold_info = parse_ns_fold_log(str(log_path)) if log_path.exists() else {}
                rows.append({
                    "prefix": prefix, "round": r, "fold": fold,
                    "status": "ckpt", "epoch": fold_info.get("epoch", "?"),
                    "cur_auc": fold_info.get("best_auc", 0),
                    "best_auc": fold_info.get("best_auc", 0),
                    "loss": fold_info.get("loss", 0),
                })
            elif log_path.exists():
                fold_info = parse_ns_fold_log(str(log_path))
                rows.append({
                    "prefix": prefix, "round": r, "fold": fold,
                    "status": "running" if fold_info["epoch"] > 0 else "started",
                    "epoch": fold_info["epoch"],
                    "cur_auc": fold_info["cur_auc"],
                    "best_auc": fold_info["best_auc"],
                    "loss": fold_info["loss"],
                })
            else:
                rows.append({
                    "prefix": prefix, "round": r, "fold": fold,
                    "status": "waiting", "epoch": 0,
                    "cur_auc": 0.0, "best_auc": 0.0, "loss": 0.0,
                })
    return rows


def parse_chain_log(name: str) -> str:
    """Return last meaningful line from an auto_*_ns_20s_full.log."""
    lines = _last_lines(LOG_DIR / name, n=20)
    if not lines:
        return "not_started"
    tag = name.split("_")[1].upper()  # SED or SSM
    last = next((l for l in reversed(lines) if f"[{tag}" in l), "")
    return last.split("]", 1)[-1].strip() if last else lines[-1].strip()


# ── Print ─────────────────────────────────────────────────────────────────────

def print_status() -> list[dict]:
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"\n{'='*68}")
    print(f"  BirdCLEF-2026 NS Chain Monitor  [{now}]")
    print(f"{'='*68}")

    # Auto chain status
    proto_chain = parse_chain_log("proto_teacher_chain.log")
    sed_chain   = parse_chain_log("auto_sed_ns_20s_full.log")
    print(f"\n  [proto-chain]   {proto_chain}")
    print(f"  [auto-sed-20s]  {sed_chain}")

    # Perch head
    ph = parse_perch_log()
    if ph["status"] == "not_started":
        print(f"\n  [perch-head]  not started")
    else:
        print(f"\n  [perch-head]  {ph['status'].upper()}")
        print(f"    ep={ph['epoch']}/{ph.get('total_ep', 80)}  "
              f"cur_auc={ph['cur_auc']:.4f}  best={ph['best_auc']:.4f}  loss={ph['loss']:.4f}")

    excel_rows = []

    clip_label = {"sed": "20s", "ssm": "10s"}
    for prefix in ["sed", "ssm"]:
        rows = parse_ns_exp(prefix)
        cl = clip_label[prefix]
        if not rows:
            print(f"\n  [{prefix}-ns-{cl}]  no data")
            continue

        print(f"\n  [{prefix.upper()} NS {cl}]")
        cur_round = None
        for row in rows:
            if row["round"] != cur_round:
                cur_round = row["round"]
                print(f"    ── Round {cur_round} ──")
            st = row["status"]
            if st == "DONE":
                print(f"      ALL FOLDS DONE  mean_auc={row['cur_auc']:.4f}")
            elif st == "waiting":
                print(f"      fold{row['fold']}  [waiting]")
            else:
                star = " ★" if st in ("ckpt",) else ""
                loss_s = row['loss'] if isinstance(row['loss'], str) else f"{row['loss']:.4f}"
                print(f"      fold{row['fold']}  [{st}]  ep={row['epoch']}  "
                      f"cur={row['cur_auc']:.4f}  best={row['best_auc']:.4f}  "
                      f"loss={loss_s}{star}")

        excel_rows.extend([{**r, "updated_at": now} for r in rows])

    # Add perch row
    excel_rows.append({
        "prefix": "perch-head", "round": 0, "fold": "all",
        "status": ph["status"], "epoch": ph.get("epoch", 0),
        "cur_auc": ph["cur_auc"], "best_auc": ph["best_auc"],
        "loss": ph["loss"], "updated_at": now,
    })

    print(f"\n{'='*68}\n")
    return excel_rows


# ── Excel ─────────────────────────────────────────────────────────────────────

def update_excel(rows: list[dict]) -> None:
    if not rows:
        return
    EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
    df_new = pd.DataFrame(rows)

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows

        if EXCEL_PATH.exists():
            wb = load_workbook(EXCEL_PATH)
        else:
            from openpyxl import Workbook
            wb = Workbook()
            wb.remove(wb.active)  # remove default sheet

        # ── "latest" sheet: overwrite each time ──────────────────────────────
        if "latest" in wb.sheetnames:
            del wb["latest"]
        ws = wb.create_sheet("latest", 0)

        # Header
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True)
        cols = list(df_new.columns)
        ws.append(cols)
        for cell in ws[1]:
            cell.fill  = header_fill
            cell.font  = header_font
            cell.alignment = Alignment(horizontal="center")

        # Rows with conditional fill
        done_fill    = PatternFill("solid", fgColor="C6EFCE")
        running_fill = PatternFill("solid", fgColor="FFEB9C")
        wait_fill    = PatternFill("solid", fgColor="F2F2F2")
        for _, row in df_new.iterrows():
            ws.append([row.get(c, "") for c in cols])
            st = str(row.get("status", ""))
            fill = done_fill if st in ("DONE", "ckpt", "done", "early_stopped") \
                   else running_fill if st in ("running", "started") \
                   else wait_fill
            for cell in ws[ws.max_row]:
                cell.fill = fill

        # Auto-width
        for col in ws.columns:
            width = max(len(str(cell.value or "")) for cell in col) + 2
            ws.column_dimensions[col[0].column_letter].width = min(width, 30)

        # ── "history" sheet: append rows each run ────────────────────────────
        if "history" not in wb.sheetnames:
            ws_h = wb.create_sheet("history")
            ws_h.append(cols)
            for cell in ws_h[1]:
                cell.fill = header_fill
                cell.font = header_font
        else:
            ws_h = wb["history"]

        for _, row in df_new.iterrows():
            ws_h.append([row.get(c, "") for c in cols])

        wb.save(EXCEL_PATH)
        print(f"  Excel updated → {EXCEL_PATH}")

    except Exception as e:
        # Fallback
        try:
            df_new.to_excel(EXCEL_PATH, index=False, sheet_name="latest")
            print(f"  Excel written (fallback) → {EXCEL_PATH}")
        except Exception as e2:
            print(f"  [Excel] failed: {e2}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", action="store_true")
    args = parser.parse_args()

    rows = print_status()
    if args.excel:
        update_excel(rows)


if __name__ == "__main__":
    main()
