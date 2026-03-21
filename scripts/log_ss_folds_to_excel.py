"""Log soundscape 4-fold experiment status to Excel (reports/exp_results.xlsx)."""
import json
import os
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

EXCEL_PATH = "reports/exp_results.xlsx"
SHEET = "SS-Folds"
FOLDS = [
    ("sed-ss-fold0", "configs/sed_ss_fold0.yaml"),
    ("sed-ss-fold1", "configs/sed_ss_fold1.yaml"),
    ("sed-ss-fold2", "configs/sed_ss_fold2.yaml"),
    ("sed-ss-fold3", "configs/sed_ss_fold3.yaml"),
]


def read_fold(name):
    p = f"outputs/{name}/result.json"
    if not os.path.exists(p):
        return {"name": name, "status": "not started", "epoch": 0, "cur_auc": 0.0,
                "best_auc": 0.0, "best_ep": 0, "total_epochs": 50, "updated": ""}
    d = json.load(open(p))
    h = d.get("epoch_history", [])
    best_auc = max((e.get("val_roc_auc", 0) for e in h), default=0.0)
    best_ep = next((e["epoch"] for e in h if e.get("val_roc_auc", 0) == best_auc), 0)
    cur_ep = h[-1]["epoch"] if h else 0
    cur_auc = h[-1].get("val_roc_auc", 0.0) if h else 0.0
    done = d.get("finished", False)
    status = "done" if done else "running"
    return {
        "name": name,
        "status": status,
        "epoch": cur_ep,
        "cur_auc": round(cur_auc, 4),
        "best_auc": round(best_auc, 4),
        "best_ep": best_ep,
        "total_epochs": d.get("total_epochs", 50),
        "updated": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }


def main():
    rows = [read_fold(name) for name, _ in FOLDS]
    df = pd.DataFrame(rows, columns=[
        "name", "status", "epoch", "cur_auc", "best_auc", "best_ep", "total_epochs", "updated"
    ])
    df.columns = ["Fold", "Status", "Epoch", "Cur AUC", "Best AUC", "Best Ep", "Total Ep", "Updated"]

    # Print to stdout for monitoring
    print(f"\n=== Soundscape 4-Fold Status [{datetime.now().strftime('%H:%M:%S')}] ===")
    for r in rows:
        flag = " 🎯" if r["best_auc"] > 0.85 else ""
        print(f"  {r['name']}: ep={r['epoch']}/{r['total_epochs']} "
              f"cur={r['cur_auc']:.4f} best={r['best_auc']:.4f}@ep{r['best_ep']} "
              f"[{r['status']}]{flag}")
    done_folds = [r for r in rows if r["status"] == "done"]
    if done_folds:
        avg = sum(r["best_auc"] for r in done_folds) / len(done_folds)
        print(f"  → {len(done_folds)}/4 done | avg best AUC = {avg:.4f}")

    # Write to Excel
    wb = load_workbook(EXCEL_PATH)
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET)

    # Header
    headers = list(df.columns)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(bold=True, color="FFFFFF")

    # Data rows
    green = PatternFill("solid", fgColor="C6EFCE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    blue = PatternFill("solid", fgColor="DDEBF7")

    for row_i, r in enumerate(rows, 2):
        vals = [r["name"], r["status"], f"{r['epoch']}/{r['total_epochs']}",
                r["cur_auc"], r["best_auc"], r["best_ep"], r["total_epochs"], r["updated"]]
        for col, val in enumerate(vals, 1):
            ws.cell(row=row_i, column=col, value=val)
        # Color by status
        fill = green if r["status"] == "done" else (blue if r["status"] == "running" else yellow)
        for col in range(1, len(vals) + 1):
            ws.cell(row=row_i, column=col).fill = fill

    # Auto column width
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)

    wb.save(EXCEL_PATH)
    print(f"  → Saved to {EXCEL_PATH} [{SHEET}]")


if __name__ == "__main__":
    os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    main()
