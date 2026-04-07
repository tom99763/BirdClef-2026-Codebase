"""Build / update report/exp_results.xlsx with all BirdCLEF-2026 experiment results.

Run at any time to refresh the Excel file from on-disk JSON artefacts.
Called by the monitoring cron job.
"""

import json
import os
import sys
from datetime import datetime

import openpyxl
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side,
                              numbers)
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUTS = os.path.join(ROOT, "outputs")
REPORT_DIR = os.path.join(ROOT, "reports")
XLSX = os.path.join(REPORT_DIR, "exp_results.xlsx")

# ── Colour palette ────────────────────────────────────────────────────────────
C_HEADER    = "1F3864"   # dark navy header
C_DONE      = "C6EFCE"   # light green
C_RUNNING   = "FFEB9C"   # light amber
C_FAILED    = "FFC7CE"   # light red
C_PLANNED   = "DDEBF7"   # light blue
C_BEST      = "00B050"   # dark green for best numbers
C_ALT_ROW   = "F2F2F2"   # alternating row

# ── Status helpers ─────────────────────────────────────────────────────────────
def read_json(path):
    try:
        with open(path) as f:
            return json.load(f)
    except Exception:
        return {}


def load_holdout(run_name):
    """Try to read holdout AUC from outputs/<run_name>/sed_holdout_eval.json."""
    p = os.path.join(OUTPUTS, run_name, "sed_holdout_eval.json")
    if os.path.exists(p):
        d = read_json(p)
        return d.get("holdout_auc")
    return None


def load_result(run_name):
    p = os.path.join(OUTPUTS, run_name, "result.json")
    return read_json(p) if os.path.exists(p) else {}


def load_geo_holdout(run_name, mask_mode):
    """Read geo-masked holdout eval JSON."""
    suffix = f"_geo-{mask_mode}"
    p = os.path.join(OUTPUTS, f"{run_name}{suffix}", "sed_holdout_eval.json")
    if os.path.exists(p):
        d = read_json(p)
        return d.get("holdout_auc")
    return None


# ── Experiment catalogue ───────────────────────────────────────────────────────
# Each entry: dict with keys:
#   name, phase, type, techniques, config, status, notes
# Dynamic fields (best_val, holdout_auc, lb_score, epoch) loaded from disk.

EXPERIMENTS = [
    # ── Phase 0: Perch / TF head ──────────────────────────────────────────────
    dict(name="nohuman-embedding-soundscape", phase="0-Perch", type="TF",
         techniques="Perch embedding + MLP head, embedding_head mode, augment=off",
         config="nohuman-embedding-soundscape.yaml",
         static_status="done", lb_score=None,
         notes="Perch embedding baseline"),
    dict(name="nohuman-label-soundscape-train", phase="0-Perch", type="TF",
         techniques="Perch label_head, train on labeled soundscapes, holdout 0.9550",
         config="exp_nohuman_label_soundscape_train.yaml",
         static_status="done", lb_score=None,
         notes="Teacher for distillation"),
    dict(name="nohuman-label-pseudo", phase="0-Perch", type="TF",
         techniques="Perch label_head + pseudo labels round5, val=0.9748",
         config="exp_nohuman_label_pseudo.yaml",
         static_status="done", lb_score=None,
         notes="Best Perch head"),

    # ── Ensemble submissions ──────────────────────────────────────────────────
    dict(name="ensemble-v5-public-trick", phase="0-Ensemble", type="Ensemble",
         techniques="Perch TFLite x3 (pseudo+soundscape+embedding) + v5 + competitor + Tricks T1/T2/T3",
         config="submissions/best_ensemble_public_trick.ipynb",
         static_status="done", lb_score=0.893,
         notes="CURRENT BEST LB. v5 holdout=0.9192."),
    dict(name="ensemble-v9-asl-soup", phase="0-Ensemble", type="Ensemble",
         techniques="Perch TFLite x3 (pseudo+soundscape+embedding) + v9-asl-soup + competitor + Tricks T1/T2/T3",
         config="submissions/v9_asl_soup_public_trick.ipynb",
         static_status="done", lb_score=0.892,
         notes="Submitted 2026-03-17. Holdout 0.9532 but LB -0.001 vs v5. Holdout≠LB proxy."),

    # ── Phase 1: SED baselines ────────────────────────────────────────────────
    dict(name="sed-b0-v3", phase="1-SED-Base", type="SED",
         techniques="EfficientNet-B0, basic SED, focal loss",
         config="sed_b0_v3.yaml",
         static_status="abandoned", lb_score=None,
         notes="Early prototype, stopped early"),
    dict(name="sed-b0-v4", phase="1-SED-Base", type="SED",
         techniques="EfficientNet-B0, SED, focal loss",
         config="sed_b0_v4.yaml",
         static_status="abandoned", lb_score=None,
         notes="Stopped early — replaced by v5"),
    dict(name="sed-b0-v5", phase="1-SED-Base", type="SED",
         techniques="B0 ns_jft_in1k, GEMFreqPool, AttentionSED, dual loss, soundscape val",
         config="sed_b0_v5.yaml",
         static_status="done", lb_score=None,
         notes="Best baseline — holdout 0.9192"),
    dict(name="sed-b0-v6", phase="1-SED-Base", type="SED",
         techniques="B0 + pseudo labels (round1)",
         config="sed_b0_v6.yaml",
         static_status="done", lb_score=None,
         notes=""),
    dict(name="sed-b0-v6-soup", phase="1-SED-Base", type="SED",
         techniques="Model soup (top-3 ckpts) of v6",
         config="sed_b0_v6.yaml",
         static_status="done", lb_score=None,
         notes=""),
    dict(name="sed-v2s-v1", phase="1-SED-Base", type="SED",
         techniques="EfficientNet-V2S backbone",
         config="sed_v2s_v1.yaml",
         static_status="done", lb_score=None,
         notes=""),
    dict(name="sed-v2s-v1-soup", phase="1-SED-Base", type="SED",
         techniques="Model soup (top-3) of V2S-v1",
         config="sed_v2s_v1.yaml",
         static_status="done", lb_score=None,
         notes=""),
    dict(name="sed-b2-v1", phase="1-SED-Base", type="SED",
         techniques="EfficientNet-B2 backbone",
         config="sed_b2_v1.yaml",
         static_status="abandoned", lb_score=None,
         notes="OOM/stopped early"),

    # ── Phase 2: Ablations ────────────────────────────────────────────────────
    dict(name="sed-b0-v9-asl", phase="2-Ablations", type="SED",
         techniques="B0 + ASL(g-=4,g+=0,clip=0.05), soundscape train, pseudo r5",
         config="sed_b0_v9_asl.yaml",
         static_status=None, lb_score=None,
         notes="ASL loss ablation"),
    dict(name="sed-b0-v10-cutmix", phase="2-Ablations", type="SED",
         techniques="B0 + CutMix augmentation, mixup=0",
         config="sed_b0_v10_cutmix.yaml",
         static_status=None, lb_score=None,
         notes="CutMix – likely failing (best@ep2)"),

    # ── Phase 3: SED-P (PCEN + dual loss + LLRD + AMP) ───────────────────────
    dict(name="sedp-b0-v1", phase="3-SEDP", type="SED-P",
         techniques="PCEN, dual loss (clip+frame), LLRD, AMP fp16",
         config="sedp_b0_v1.yaml",
         static_status="planned", lb_score=None,
         notes=""),
    dict(name="sedp-b0-v2-fusion", phase="3-SEDP", type="SED-P",
         techniques="PCEN + dual + LLRD + AMP + ASL + trainable PCEN",
         config="sedp_b0_v2_fusion.yaml",
         static_status="planned", lb_score=None,
         notes="Full fusion — primary experiment"),
    dict(name="sedp-b0-v3-abl-no-pcen", phase="3-SEDP", type="SED-P",
         techniques="SEDP-v2 without PCEN (ablation)",
         config="sedp_b0_v3_abl_no_pcen.yaml",
         static_status="planned", lb_score=None,
         notes="PCEN ablation"),

    # ── Phase 4: Geographic filtering ────────────────────────────────────────
    dict(name="geo-ss_soft-sf0.10", phase="4-GeoFilter", type="Post-proc",
         techniques="ss_soft mask: soundscape→1.0, others→0.1 (post-process)",
         config="scripts/run_geo.sh",
         static_status="planned", lb_score=None,
         notes="Applied to v5, v9, SEDP checkpoints"),
    dict(name="geo-ss_hard", phase="4-GeoFilter", type="Post-proc",
         techniques="ss_hard mask: soundscape→1.0, others→0.0",
         config="scripts/run_geo.sh",
         static_status="planned", lb_score=None,
         notes="Aggressive; 75/234 species kept"),
    dict(name="geo-sa_weighted", phase="4-GeoFilter", type="Post-proc",
         techniques="sa_weighted: pred *= max(in_soundscape, sa_fraction)",
         config="scripts/run_geo.sh",
         static_status="planned", lb_score=None,
         notes="Continuous — 204/234 with SA recs"),

    # ── Phase 5: Distillation ─────────────────────────────────────────────────
    dict(name="distill-b0-v1", phase="5-Distill", type="SED-Distill",
         techniques="Perch teacher (1176 clips), B0 student, distill_w=0.5",
         config="distill_b0_v1.yaml",
         static_status="planned", lb_score=None,
         notes="Uses existing pseudo_labels/round5_pseudo.csv"),
    dict(name="distill-b0-v2-full", phase="5-Distill", type="SED-Distill",
         techniques="Perch teacher (ALL 10658 soundscapes ~256K clips), distill_w=0.5",
         config="distill_b0_v2_full.yaml",
         static_status="planned", lb_score=None,
         notes="Requires extract_perch_teacher_all_ss.py"),

    # ── Phase 6: Round 2 ablations ────────────────────────────────────────────
    dict(name="sed-b0-v11-soft-sec", phase="6-Round2", type="SED",
         techniques="MaskedBCE: secondary labels as 0.5, mask gradient",
         config="sed_b0_v11_soft_sec.yaml",
         static_status="planned", lb_score=None, notes=""),
    dict(name="sed-b0-v12-bce", phase="6-Round2", type="SED",
         techniques="Plain BCE loss (vs ASL ablation)",
         config="sed_b0_v12_bce.yaml",
         static_status="planned", lb_score=None, notes=""),
    dict(name="sed-b0-v13-asl-cutmix", phase="6-Round2", type="SED",
         techniques="ASL + CutMix combined",
         config="sed_b0_v13_asl_cutmix.yaml",
         static_status="planned", lb_score=None, notes=""),
    dict(name="sed-b0-v14-50ep", phase="6-Round2", type="SED",
         techniques="50 epochs (vs 30 baseline)",
         config="sed_b0_v14_50ep.yaml",
         static_status="planned", lb_score=None, notes=""),
    dict(name="sed-b0-v15-no-sec", phase="6-Round2", type="SED",
         techniques="No secondary labels, in_chans=1 (matches embed-distill-b0-v1 backbone)",
         config="sed_b0_v15_no_sec.yaml",
         static_status=None, lb_score=None,
         notes="Warm-started from embed-distill-b0-v1 backbone"),
    dict(name="sed-b0-v15-no-sec-soup", phase="6-Round2", type="SED",
         techniques="Model soup of v15-no-sec top checkpoints",
         config="sed_b0_v15_no_sec.yaml",
         static_status=None, lb_score=None,
         notes="Soup of distill-backbone v15"),
    dict(name="sed-b0-v16-rating3", phase="6-Round2", type="SED",
         techniques="min_rating=3.0 (quality filter), warm-start from best embed-distill backbone",
         config="sed_b0_v16_rating3.yaml",
         static_status=None, lb_score=None,
         notes="Chain3: uses best of distill-b0-v2/v3/v4 backbone"),

    # ── Phase 7: Round 3 dual-loss sweeps ─────────────────────────────────────
    dict(name="sed-b0-v17-dual30", phase="7-Round3", type="SED",
         techniques="Dual loss, 30 ep, standard aug; distill backbone",
         config="sed_b0_v17_dual30.yaml",
         static_status=None, lb_score=None,
         notes="Chain3: best distill backbone"),
    dict(name="sed-b0-v17-dual30-soup", phase="7-Round3", type="SED",
         techniques="Model soup of v17-dual30",
         config="sed_b0_v17_dual30.yaml",
         static_status=None, lb_score=None, notes="Chain3 soup"),
    dict(name="sed-b0-v18-dual-ss10", phase="7-Round3", type="SED",
         techniques="Dual loss + soundscape oversample x10; distill backbone",
         config="sed_b0_v18_dual_ss10.yaml",
         static_status=None, lb_score=None,
         notes="Chain3: best distill backbone"),
    dict(name="sed-b0-v18-dual-ss10-soup", phase="7-Round3", type="SED",
         techniques="Model soup of v18-dual-ss10",
         config="sed_b0_v18_dual_ss10.yaml",
         static_status=None, lb_score=None, notes="Chain3 soup"),
    dict(name="sed-b0-v19-dual-freqmask", phase="7-Round3", type="SED",
         techniques="Dual loss + stronger freq masking; distill backbone",
         config="sed_b0_v19_dual_freqmask.yaml",
         static_status=None, lb_score=None,
         notes="Chain3: best distill backbone"),
    dict(name="sed-b0-v19-dual-freqmask-soup", phase="7-Round3", type="SED",
         techniques="Model soup of v19-dual-freqmask",
         config="sed_b0_v19_dual_freqmask.yaml",
         static_status=None, lb_score=None, notes="Chain3 soup"),
    dict(name="sed-b0-v20-pseudo-r5", phase="7-Round3", type="SED",
         techniques="v15 base + Perch round5 pseudo labels (1176 soundscape clips, threshold=0.2) as extra training data",
         config="configs/sed_b0_v20_pseudo_r5.yaml",
         static_status=None, lb_score=None,
         notes="Chain3: Perch pseudo label soundscape augmentation"),
    dict(name="sed-b0-v20-pseudo-r5-soup", phase="7-Round3", type="SED",
         techniques="Model soup of v20-pseudo-r5",
         config="configs/sed_b0_v20_pseudo_r5.yaml",
         static_status=None, lb_score=None, notes="Chain3 soup"),
    dict(name="sed-b0-v20-dual-mixup08", phase="7-Round3", type="SED",
         techniques="Dual loss + mixup alpha=0.8",
         config="sed_b0_v20_dual_mixup08.yaml",
         static_status="planned", lb_score=None, notes=""),
    dict(name="sed-b0-v21-dual-rating3", phase="7-Round3", type="SED",
         techniques="Dual loss + min_rating=3.0",
         config="sed_b0_v21_dual_rating3.yaml",
         static_status="planned", lb_score=None, notes=""),
    dict(name="sed-b0-v22-dual-noclipmix", phase="7-Round3", type="SED",
         techniques="Dual loss, no clip mixup",
         config="sed_b0_v22_dual_noclipmix.yaml",
         static_status=None, lb_score=None,
         notes="ClipMix ablation — best 0.7485; proved ClipMix is necessary"),

    # ── Phase 7b: B0 Round4 kitchen-sink sweeps ───────────────────────────────
    dict(name="sed-b0-v23-perch-asl-aug", phase="7b-Round4", type="SED",
         techniques="B0 + ASL + Perch soft-pseudo aug, heavy augmentation",
         config="sed_b0_v23.yaml",
         static_status=None, lb_score=None,
         notes="Best val 0.7367 @ep4 — aug too heavy"),
    dict(name="sed-b0-v24-soft-pseudo", phase="7b-Round4", type="SED",
         techniques="B0 + ASL + soft pseudo (r5) + soundscape oversample=10, dual loss",
         config="sed_b0_v24.yaml",
         static_status=None, lb_score=None,
         notes="Same as v30-multipseu config; holdout 0.9199"),
    dict(name="sed-b0-v26-asl-npcen", phase="7b-Round4", type="SED",
         techniques="B0 + ASL + NPCEN frontend",
         config="sed_b0_v26.yaml",
         static_status=None, lb_score=None,
         notes="Val 0.7262 @ep3; holdout 0.9574"),
    dict(name="sed-b0-v27-soft-boost", phase="7b-Round4", type="SED",
         techniques="B0 + ASL + soft pseudo boost (oversample=20), multi-round pseudo",
         config="sed_b0_v27.yaml",
         static_status=None, lb_score=None,
         notes="Best val 0.7804 @ep10; holdout 0.9915 🎯"),
    dict(name="sed-b0-v28-final-combo", phase="7b-Round4", type="SED",
         techniques="B0 + ASL + dual loss + soft KD (10x r5) + multi-round pseudo (r1-r5) + SpecAugment",
         config="sed_b0_v28.yaml",
         static_status=None, lb_score=None,
         notes="Kitchen sink formula; val 0.8053 @ep9; holdout 0.9861 🎯"),

    # ── Phase 7c: V2S experiments ─────────────────────────────────────────────
    dict(name="sed-v2s-v2-asl", phase="7c-V2S", type="SED",
         techniques="EfficientNetV2-S + ASL, soundscape fine-tune",
         config="sed_v2s_v2.yaml",
         static_status=None, lb_score=None,
         notes="Val 0.7113 @ep4; holdout 0.9717 — V2S generalises well despite low val"),
    dict(name="sed-v2s-v3-full-kitchen", phase="7c-V2S", type="SED",
         techniques="EfficientNetV2-S + full v28 formula (ASL+dual+soft KD+multi-pseudo)",
         config="sed_v2s_v3.yaml",
         static_status=None, lb_score=None,
         notes="Val 0.7592 @ep5; early stop ep11. Full formula doesn't transfer from B0 to V2S"),

    # ── Phase 7d: B0 multi-pseudo / background noise ──────────────────────────
    dict(name="sed-b0-v30-multipseu", phase="7d-MultiPseudo", type="SED",
         techniques="B0 + ASL + multi-round pseudo (r1-r5) + soft KD (10x r5) + SpecAugment",
         config="sed_b0_v30_multipseu.yaml",
         static_status=None, lb_score=None,
         notes="NEW SOTA: val 0.8139 @ep6; holdout 0.9839 🎯"),
    dict(name="sed-b0-v30-bgnoise", phase="7d-MultiPseudo", type="SED",
         techniques="v30-multipseu + background noise augmentation (SNR=[5,30]dB)",
         config="sed_b0_v30_bgnoise.yaml",
         static_status=None, lb_score=None,
         notes="Val 0.7725 @ep4; holdout 0.9627. BgNoise hurt soundscape val (-0.04)"),

    # ── Phase 7e: EfficientNet-B3 experiments ─────────────────────────────────
    dict(name="sed-b3-v1-asl", phase="7e-B3", type="SED",
         techniques="EfficientNet-B3 NS-JFT + ASL (g-=4,g+=0,clip=0.05) + multi-round pseudo",
         config="sed_b3_v1.yaml",
         static_status=None, lb_score=None,
         notes="Val 0.7805 @ep2; holdout 0.9553 🎯. Large gap: B3 generalises despite unstable val"),
    dict(name="sed-b3-v1-fold0", phase="7e-B3", type="SED",
         techniques="B3 v1-asl, 4-fold CV fold0 (~26661 recordings)",
         config="sed_b3_v1_fold0.yaml",
         static_status=None, lb_score=None,
         notes="CV fold0; ep1-ep5 trend: 0.7525→0.7441→0.7532→0.7678→0.7778"),
    dict(name="sed-b3-v1-fold1", phase="7e-B3", type="SED",
         techniques="B3 v1-asl, 4-fold CV fold1 (~26662 recordings)",
         config="sed_b3_v1_fold1.yaml",
         static_status=None, lb_score=None,
         notes="CV fold1; ep1-ep3: 0.7424→0.7184→0.7260 (recovering)"),

    # ── Phase 8: Embedding-level Distillation (B0 backbone → Perch emb space) ──
    dict(name="embed-distill-b0-v1", phase="8-EmbDistill", type="Embed-Distill",
         techniques="B0 in_chans=1, cosine loss, SpecAugment×1 (freq=30,time=48), epochs=30",
         config="configs/embed_distill_b0_v1.yaml",
         static_status=None, lb_score=None,
         notes="Baseline embed distill; best_val_cos=0.6202@ep17, early stop ep22"),
    dict(name="embed-distill-b0-v2", phase="8-EmbDistill", type="Embed-Distill",
         techniques="B0 in_chans=3, cosine loss, SpecAugment×1 (freq=30,time=48), epochs=30",
         config="configs/embed_distill_b0_v2.yaml",
         static_status=None, lb_score=None,
         notes="Tests in_chans=3 (ImageNet pretrain convention)"),
    dict(name="embed-distill-b0-v3", phase="8-EmbDistill", type="Embed-Distill",
         techniques="B0 in_chans=1, InfoNCE loss (NT-Xent temp=0.1), SpecAugment×1, epochs=30",
         config="configs/embed_distill_b0_v3.yaml",
         static_status=None, lb_score=None,
         notes="Contrastive loss: batch negatives → discriminative representations"),
    dict(name="embed-distill-b0-v4", phase="8-EmbDistill", type="Embed-Distill",
         techniques="B0 in_chans=1, cosine loss, SpecAugment×2 (freq=40×2,time=64×2), Mixup(0.4), Noise(0.05), epochs=30",
         config="configs/embed_distill_b0_v4.yaml",
         static_status=None, lb_score=None,
         notes="Heavy augmentation: double masks + mixup + gaussian noise"),

    # ── Phase 9: Backbone Comparison (distill → frozen → SED head) ────────────
    dict(name="embed-distill-b2-v1", phase="9-BackboneCmp", type="Embed-Distill",
         techniques="B2 (7.7M) in_chans=1, cosine loss, SpecAugment×1, epochs=30",
         config="configs/embed_distill_b2_v1.yaml",
         static_status=None, lb_score=None,
         notes="Chain4: distill B2 backbone before head-only SED"),
    dict(name="sed-b2-v1-distill-head", phase="9-BackboneCmp", type="SED",
         techniques="B2 frozen backbone (embed-distill-b2-v1), head-only training",
         config="configs/sed_b2_v1_distill_head.yaml",
         static_status=None, lb_score=None,
         notes="Chain4: head-only SED with frozen B2"),
    dict(name="sed-b2-v1-distill-head-soup", phase="9-BackboneCmp", type="SED",
         techniques="Model soup of sed-b2-v1-distill-head",
         config="configs/sed_b2_v1_distill_head.yaml",
         static_status=None, lb_score=None, notes="Chain4 soup"),
    dict(name="embed-distill-b4-v1", phase="9-BackboneCmp", type="Embed-Distill",
         techniques="B4 (17.5M) in_chans=1, cosine loss, SpecAugment×1, epochs=30",
         config="configs/embed_distill_b4_v1.yaml",
         static_status=None, lb_score=None,
         notes="Chain4: distill B4 backbone before head-only SED"),
    dict(name="sed-b4-v1-distill-head", phase="9-BackboneCmp", type="SED",
         techniques="B4 frozen backbone (embed-distill-b4-v1), head-only training",
         config="configs/sed_b4_v1_distill_head.yaml",
         static_status=None, lb_score=None,
         notes="Chain4: head-only SED with frozen B4"),
    dict(name="sed-b4-v1-distill-head-soup", phase="9-BackboneCmp", type="SED",
         techniques="Model soup of sed-b4-v1-distill-head",
         config="configs/sed_b4_v1_distill_head.yaml",
         static_status=None, lb_score=None, notes="Chain4 soup"),
    dict(name="embed-distill-convnext-v1", phase="9-BackboneCmp", type="Embed-Distill",
         techniques="ConvNeXt-Tiny (28M) in_chans=1, cosine loss, SpecAugment×1, epochs=30",
         config="configs/embed_distill_convnext_v1.yaml",
         static_status=None, lb_score=None,
         notes="Chain4: distill ConvNeXt-Tiny backbone"),
    dict(name="sed-convnext-v1-distill-head", phase="9-BackboneCmp", type="SED",
         techniques="ConvNeXt-Tiny frozen backbone (embed-distill-convnext-v1), head-only",
         config="configs/sed_convnext_v1_distill_head.yaml",
         static_status=None, lb_score=None,
         notes="Chain4: head-only SED with frozen ConvNeXt-Tiny"),
    dict(name="sed-convnext-v1-distill-head-soup", phase="9-BackboneCmp", type="SED",
         techniques="Model soup of sed-convnext-v1-distill-head",
         config="configs/sed_convnext_v1_distill_head.yaml",
         static_status=None, lb_score=None, notes="Chain4 soup"),
    dict(name="embed-distill-convnext-small-v1", phase="9-BackboneCmp", type="Embed-Distill",
         techniques="ConvNeXt-Small (49.5M) in_chans=1, cosine loss, SpecAugment×1, batch=48, epochs=30",
         config="configs/embed_distill_convnext_small_v1.yaml",
         static_status=None, lb_score=None,
         notes="Chain4: larger ConvNeXt vs Tiny comparison"),
    dict(name="sed-convnext-small-v1-distill-head", phase="9-BackboneCmp", type="SED",
         techniques="ConvNeXt-Small frozen backbone, head-only training, batch=24",
         config="configs/sed_convnext_small_v1_distill_head.yaml",
         static_status=None, lb_score=None,
         notes="Chain4: head-only SED with frozen ConvNeXt-Small"),
    dict(name="sed-convnext-small-v1-distill-head-soup", phase="9-BackboneCmp", type="SED",
         techniques="Model soup of sed-convnext-small-v1-distill-head",
         config="configs/sed_convnext_small_v1_distill_head.yaml",
         static_status=None, lb_score=None, notes="Chain4 soup"),

    # ── Phase 10: SED Improvement Ablations (v31-v36) ────────────────────────
    dict(name="sed-b0-v31-lr1e3", phase="10-SEDImprove", type="SED",
         techniques="BCE dual, cosine no-warmup, lr=1e-3, soundscape_val_frac=1.0",
         config="configs/sed_b0_v31_lr1e3.yaml",
         static_status="abandoned", lb_score=None,
         notes="Skipped — jumped directly to v33"),
    dict(name="sed-b0-v32-lr5e4", phase="10-SEDImprove", type="SED",
         techniques="BCE dual, cosine no-warmup, lr=5e-4 (ablation vs v5 warmup)",
         config="configs/sed_b0_v32_lr5e4.yaml",
         static_status="abandoned", lb_score=None,
         notes="Skipped — jumped directly to v33"),
    dict(name="sed-b0-v33-warmrestart", phase="10-SEDImprove", type="SED",
         techniques="BCE dual, CosineWarmRestarts T0=10, lr=1e-3, no-warmup, val_frac=1.0",
         config="configs/sed_b0_v33_warmrestart.yaml",
         static_status=None, lb_score=None,
         notes="Scheduler ablation: periodic LR resets every 10 epochs"),
    dict(name="sed-b0-v34-focal-g2", phase="10-SEDImprove", type="SED",
         techniques="FocalBCE γ=2.0, α=0.75 (EDA-corrected), lr=1e-3, no-warmup, val_frac=1.0",
         config="configs/sed_b0_v34_focal_g2.yaml",
         static_status=None, lb_score=None,
         notes="Focal standard. EDA: median neg:pos=341:1. alpha=0.75 avoids trivial min"),
    dict(name="sed-b0-v35-focal-g3", phase="10-SEDImprove", type="SED",
         techniques="FocalBCE γ=3.0, α=0.75 (higher focus for extreme 35k:1 classes), lr=1e-3",
         config="configs/sed_b0_v35_focal_g3.yaml",
         static_status=None, lb_score=None,
         notes="Ablation vs v34: gamma 2→3 for 53 zero-sample classes"),
    dict(name="sed-b0-v36-pos-weight", phase="10-SEDImprove", type="SED",
         techniques="BCEPosWeight sqrt(n_neg/n_pos) clip=20, lr=1e-3, no-warmup, val_frac=1.0",
         config="configs/sed_b0_v36_pos_weight.yaml",
         static_status=None, lb_score=None,
         notes="Per-class fixed weights vs focal sample-adaptive weights"),
]


# ── Derive live status from disk ──────────────────────────────────────────────
def get_status(exp):
    if exp.get("static_status") in ("done", "abandoned", "planned"):
        return exp["static_status"]

    r = load_result(exp["name"])
    if not r:
        return "planned"
    if r.get("finished"):
        return "done"
    # embed-distill uses best_val_cos, SED uses total_epochs_run
    if r.get("total_epochs_run", 0) > 0 or r.get("best_val_cos") is not None:
        return "running"
    return "planned"


def get_best_val(exp):
    r = load_result(exp["name"])
    # Embed-distill experiments use best_val_cos instead of best_val_roc_auc
    if exp.get("type") == "Embed-Distill":
        return r.get("best_val_cos")
    return r.get("best_val_roc_auc")


def get_current_epoch(exp):
    r = load_result(exp["name"])
    epochs = r.get("epoch_history", [])
    if epochs:
        return epochs[-1].get("epoch")
    # embed-distill stores current epoch directly
    ep = r.get("epoch") or r.get("total_epochs_run")
    return ep


def get_holdout(exp):
    h = load_holdout(exp["name"])
    if h:
        return h
    # Also try result.json key (for TF models)
    r = load_result(exp["name"])
    return r.get("holdout_auc")


# ── Excel builder ─────────────────────────────────────────────────────────────
STATUS_FILL = {
    "done":     PatternFill("solid", fgColor=C_DONE),
    "running":  PatternFill("solid", fgColor=C_RUNNING),
    "failed":   PatternFill("solid", fgColor=C_FAILED),
    "abandoned":PatternFill("solid", fgColor=C_FAILED),
    "planned":  PatternFill("solid", fgColor=C_PLANNED),
}

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def hdr_style(cell, text):
    cell.value = text
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor=C_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def build_xlsx():
    os.makedirs(REPORT_DIR, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Experiments"

    # ── Header row ─────────────────────────────────────────────────────────────
    COLS = [
        "#", "Phase", "Experiment Name", "Type",
        "Key Techniques",
        "Status", "Epoch",
        "Val Metric", "Holdout AUC", "LB Score",
        "Config", "Notes", "Updated",
    ]
    for ci, h in enumerate(COLS, 1):
        hdr_style(ws.cell(1, ci), h)

    ws.row_dimensions[1].height = 32

    # ── Column widths ─────────────────────────────────────────────────────────
    widths = [4, 13, 28, 10, 58, 10, 7, 12, 12, 10, 30, 35, 18]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Data rows ─────────────────────────────────────────────────────────────
    best_holdout = max(
        (get_holdout(e) or 0) for e in EXPERIMENTS
    )

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    for ri, exp in enumerate(EXPERIMENTS, 2):
        status    = get_status(exp)
        best_val  = get_best_val(exp)
        holdout   = get_holdout(exp)
        cur_epoch = get_current_epoch(exp)

        row_fill = PatternFill("solid", fgColor=C_ALT_ROW) if ri % 2 == 0 else None

        vals = [
            ri - 1,
            exp["phase"],
            exp["name"],
            exp["type"],
            exp["techniques"],
            status,
            cur_epoch,
            best_val,
            holdout,
            exp.get("lb_score"),
            exp["config"],
            exp.get("notes", ""),
            now_str,
        ]

        for ci, v in enumerate(vals, 1):
            cell = ws.cell(ri, ci)
            cell.value = v
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(ci in (5, 12)))
            if ci == 6:
                cell.fill = STATUS_FILL.get(status, PatternFill())
                cell.font = Font(bold=True, size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif row_fill:
                cell.fill = row_fill

            # Highlight best holdout
            if ci == 9 and holdout and abs(holdout - best_holdout) < 1e-6 and holdout > 0:
                cell.font = Font(bold=True, color=C_BEST)

            # Format floats
            if ci in (8, 9, 10) and isinstance(v, float):
                cell.number_format = "0.0000"

    # ── Freeze pane + filter ──────────────────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 16

    summary_data = [
        ("Generated", now_str),
        ("Total Experiments", len(EXPERIMENTS)),
        ("Done", sum(1 for e in EXPERIMENTS if get_status(e) == "done")),
        ("Running", sum(1 for e in EXPERIMENTS if get_status(e) == "running")),
        ("Planned", sum(1 for e in EXPERIMENTS if get_status(e) == "planned")),
        ("Abandoned", sum(1 for e in EXPERIMENTS if get_status(e) == "abandoned")),
        ("", ""),
        ("Best Holdout AUC", best_holdout if best_holdout else "—"),
        ("Competitor Reference LB", 0.9883),
        ("Our Best LB", "—"),
    ]

    for ri, (k, v) in enumerate(summary_data, 1):
        ws2.cell(ri, 1).value = k
        ws2.cell(ri, 1).font = Font(bold=True)
        ws2.cell(ri, 2).value = v
        if isinstance(v, float):
            ws2.cell(ri, 2).number_format = "0.0000"

    wb.save(XLSX)
    print(f"[update_exp_results] Saved → {XLSX}  ({now_str})")
    return XLSX


if __name__ == "__main__":
    build_xlsx()
