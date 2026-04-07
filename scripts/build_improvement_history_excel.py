"""
Build improvement history Excel report for embed prior methods.
Output: reports/embed_prior_improvement_history.xlsx
"""
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
import os
os.chdir("/home/lab/BirdClef-2026-Codebase")

wb = openpyxl.Workbook()

# ── Colors ─────────────────────────────────────────────────────────────────────
C_HEADER     = "1F3864"   # dark navy
C_STAGE1     = "D6E4F7"   # light blue  — KNN stage
C_STAGE2     = "D5E8D4"   # light green — Bridge stage
C_STAGE3     = "FFE6CC"   # light orange — KDE stage
C_BEST       = "FF0000"   # red text for best rows
C_BREAK      = "FFF2CC"   # yellow — breakthrough row
C_GRAY       = "F2F2F2"
C_WHITE      = "FFFFFF"

def header_style(cell, text, bg=C_HEADER):
    cell.value = text
    cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"))

def data_style(cell, text, bold=False, color="000000", bg=C_WHITE, align="left", wrap=False):
    cell.value = text
    cell.font = Font(name="Calibri", bold=bold, color=color, size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"))

def num_style(cell, val, fmt="0.0000", bold=False, color="000000", bg=C_WHITE):
    cell.value = val
    cell.number_format = fmt
    cell.font = Font(name="Calibri", bold=bold, color=color, size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"))

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 1: 完整改進歷程 (Full Improvement History)
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "改進歷程"

ws1.column_dimensions['A'].width = 6
ws1.column_dimensions['B'].width = 20
ws1.column_dimensions['C'].width = 38
ws1.column_dimensions['D'].width = 14
ws1.column_dimensions['E'].width = 14
ws1.column_dimensions['F'].width = 10
ws1.column_dimensions['G'].width = 42

ws1.row_dimensions[1].height = 30
ws1.row_dimensions[2].height = 18

# Title
ws1.merge_cells('A1:G1')
title_cell = ws1['A1']
title_cell.value = "BirdCLEF 2026 — Embed Prior 改進歷程"
title_cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
title_cell.fill = PatternFill("solid", fgColor=C_HEADER)
title_cell.alignment = Alignment(horizontal="center", vertical="center")

# Column headers
headers = ["#", "階段 / Stage", "方法名稱", "EP-only AUC", "全管線 AUC", "進步幅度", "關鍵說明"]
for col, h in enumerate(headers, 1):
    header_style(ws1.cell(row=2, column=col), h)

# Data rows
rows = [
    # (row_num, stage_label, method, ep_only, full_pipe, delta, notes, bg_color, is_breakthrough, is_best)
    # ── VLOM Baseline ──
    (1,  "Baseline",      "VLOM（ProtoSSM + SED 50/50）",              None,  0.8137, None,   "ProtoSSM v2 + SED noisy-student r1 混合基準", C_GRAY, False, False),
    (2,  "Baseline",      "v7-geo-knn（地理 + KNN embed prior）",       None,  0.9246, 0.1109, "加入地理先驗 + KNN 嵌入先驗；首次突破 0.92",  C_GRAY, True,  False),
    # ── Stage 1: KNN ──
    (3,  "Stage 1: KNN",  "distance_weighted_knn",                      0.841, None,   None,   "最基礎距離加權 KNN embed prior",              C_STAGE1, False, False),
    (4,  "Stage 1: KNN",  "logit_max_knn（K 值掃描）",                  0.892, None,   0.051,  "最大 logit 加權 KNN，K sweep 最佳化",         C_STAGE1, False, False),
    (5,  "Stage 1: KNN",  "per_species_alpha_knn3",                     0.903, None,   0.011,  "每物種獨立混合權重 α",                        C_STAGE1, False, False),
    (6,  "Stage 1: KNN",  "interaction_knn（交互特徵）",                 0.920, 0.9412, 0.017,  "KNN EP-only 天花板；全管線提升有限",          C_STAGE1, True,  False),
    # ── Stage 2: Bridge ──
    (7,  "Stage 2: Bridge", "SS Bridge（127,896 soundscape 視窗）",      None,  0.9440, 0.0028, "用所有測試 SS 視窗作橋樑，轉導學習",          C_STAGE2, True,  False),
    (8,  "Stage 2: Bridge", "SED-Species Bridge（物種加權橋樑）",        None,  0.9444, 0.0004, "perch_sim × (1 + β × max_SED) 加權；Bridge 天花板", C_STAGE2, False, False),
    # ── Stage 3: KDE ──
    (9,  "Stage 3: KDE",  "kde_per_species（檔案級 KDE）",               0.937, 0.9560, 0.0116, "PCA-32 空間 Gaussian KDE；從相似度→密度估計的突破", C_STAGE3, True,  False),
    (10, "Stage 3: KDE",  "kde_window_level（739 視窗級 KDE）",          None,  0.9701, 0.0141, "739 訓練視窗替代 66 檔案平均；正樣本更豐富（avg 76.2 視窗/物種）", C_STAGE3, True,  False),
    (11, "Stage 3: KDE",  "kde_win_rknn_blend（KDE + RKNN k5）",        None,  0.9711, 0.0010, "密度估計 + 互惠近鄰；35%KDE + 65%RKNN，a=0.92, b=1.4", C_STAGE3, False, False),
    (12, "Stage 3: KDE",  "kde_perwin（每視窗單獨計算 → 平均）",         None,  0.9721, 0.0010, "每視窗 KDE(x_i) → avg；保留視窗內異質性；a=0.90, b=2.0 ← 目前最佳", C_STAGE3, False, True),
]

for i, (num, stage, method, ep_only, full_pipe, delta, notes, bg, is_break, is_best) in enumerate(rows, start=3):
    ws1.row_dimensions[i].height = 22

    # Stage label — merge across same-stage rows
    text_color = C_BEST if is_best else ("000000" if not is_break else "7B3F00")
    bold = is_best or is_break

    data_style(ws1.cell(row=i, column=1), num, align="center", bg=bg)
    data_style(ws1.cell(row=i, column=2), stage, bold=is_break, bg=bg, wrap=True)
    data_style(ws1.cell(row=i, column=3), method, bold=bold, color=text_color, bg=bg)

    if ep_only is not None:
        num_style(ws1.cell(row=i, column=4), ep_only, bold=bold, bg=bg)
    else:
        data_style(ws1.cell(row=i, column=4), "—", align="center", bg=bg)

    if full_pipe is not None:
        num_style(ws1.cell(row=i, column=5), full_pipe, bold=bold,
                  color=C_BEST if is_best else "000000", bg=bg)
    else:
        data_style(ws1.cell(row=i, column=5), "—", align="center", bg=bg)

    if delta is not None:
        sign = "+" if delta >= 0 else ""
        data_style(ws1.cell(row=i, column=6), f"{sign}{delta:.4f}",
                   bold=bold, align="center",
                   color="2E7D32" if delta >= 0.005 else ("1565C0" if delta > 0 else "000000"),
                   bg=bg)
    else:
        data_style(ws1.cell(row=i, column=6), "—", align="center", bg=bg)

    data_style(ws1.cell(row=i, column=7), notes, bg=bg, wrap=True)

# Freeze panes
ws1.freeze_panes = "A3"

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 2: KDE 詳細分析 (KDE Deep Dive)
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("KDE 詳細分析")

ws2.column_dimensions['A'].width = 30
ws2.column_dimensions['B'].width = 14
ws2.column_dimensions['C'].width = 14
ws2.column_dimensions['D'].width = 14
ws2.column_dimensions['E'].width = 14
ws2.column_dimensions['F'].width = 14
ws2.column_dimensions['G'].width = 40

ws2.merge_cells('A1:G1')
t2 = ws2['A1']
t2.value = "KDE 方法詳細比較（含超參數）"
t2.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
t2.fill = PatternFill("solid", fgColor="1A5276")
t2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 28

headers2 = ["方法", "PCA 維度", "頻寬 BW", "係數 a", "係數 b", "全管線 AUC", "備注"]
for col, h in enumerate(headers2, 1):
    header_style(ws2.cell(row=2, column=col), h, bg="1A5276")
ws2.row_dimensions[2].height = 18

kde_rows = [
    ("kde_per_species（檔案平均後 KDE）",    32, 1.0, 0.95, 1.2,  0.9560, "baseline: avg file emb → PCA → KDE"),
    ("embed_prior_kde_file（bw sweep）",     32, 0.5, 0.95, 1.2,  0.9560, "file-level, validated"),
    ("kde_window_level（視窗平均後計算）",   32, 0.5, 0.90, 1.2,  0.9701, "avg windows → KDE; win_k1 blend wg=0.30"),
    ("kde_pca16",                            16, 0.5, 0.90, 1.2,  0.9695, "less PCA dims"),
    ("adaptive_bandwidth_kde",               32, "適應",0.90,1.2, 0.9688, "per-species adaptive bw"),
    ("kde_win + RKNN k3",                    32, 0.5, 0.95, 1.4,  0.9709, "RKNN k=3 替代 win_k1"),
    ("kde_win + RKNN k5 (best blend)",       32, 0.5, 0.92, 1.4,  0.9711, "0.35×KDE + 0.65×RKNN, no win_k1"),
    ("kde_win + RKNN k7",                    32, 0.5, 0.92, 1.4,  0.9708, "RKNN k=7"),
    ("kde_win bw=0.6 + RKNN k5",            32, 0.6, 0.95, 1.2,  0.9707, "bw=0.6 KDE"),
    ("kde_win pca48",                        48, 0.5, 0.95, 1.4,  0.9665, "more PCA dims, worse"),
    ("kde_win pca64",                        64, 0.5, 0.95, 1.2,  0.9676, "more PCA dims, worse"),
    ("kde_perwin bw=0.4",                    32, 0.4, 0.90, 2.0,  0.9721, "per-window, bw=0.4, tied best"),
    ("kde_perwin bw=0.5 ← 目前最佳",        32, 0.5, 0.90, 2.0,  0.9721, "per-window scoring then avg"),
    ("kde_perwin bw=0.6",                    32, 0.6, 0.95, 1.8,  0.9715, "per-window, bw=0.6"),
    ("kde_perwin + RKNN k5",                 32, 0.5, 0.90, 1.4,  0.9714, "per-window + RKNN blend"),
    ("vmf_kde + RKNN k5",                    "—","—", 0.90, 1.8,  0.9607, "von Mises-Fisher kernel (cosine)"),
]

for i, row in enumerate(kde_rows, start=3):
    method, pca_n, bw, a, b, auc, note = row
    ws2.row_dimensions[i].height = 20
    is_best = auc >= 0.9721
    bg = "FFF3E0" if is_best else (C_GRAY if auc < 0.97 else C_WHITE)
    bold = is_best

    data_style(ws2.cell(row=i, column=1), method, bold=bold,
               color=C_BEST if is_best else "000000", bg=bg, wrap=True)
    data_style(ws2.cell(row=i, column=2), str(pca_n), align="center", bg=bg, bold=bold)
    data_style(ws2.cell(row=i, column=3), str(bw), align="center", bg=bg, bold=bold)
    data_style(ws2.cell(row=i, column=4), str(a), align="center", bg=bg, bold=bold)
    data_style(ws2.cell(row=i, column=5), str(b), align="center", bg=bg, bold=bold)
    num_style(ws2.cell(row=i, column=6), auc, bold=bold,
              color=C_BEST if is_best else "000000", bg=bg)
    data_style(ws2.cell(row=i, column=7), note, bg=bg, wrap=True)

ws2.freeze_panes = "A3"

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 3: 方法論對比 (Method Comparison)
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("方法論對比")

ws3.column_dimensions['A'].width = 22
ws3.column_dimensions['B'].width = 16
ws3.column_dimensions['C'].width = 16
ws3.column_dimensions['D'].width = 16
ws3.column_dimensions['E'].width = 40

ws3.merge_cells('A1:E1')
t3 = ws3['A1']
t3.value = "三大方法論對比：KNN vs Bridge vs KDE"
t3.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
t3.fill = PatternFill("solid", fgColor="1B4F72")
t3.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 28

headers3 = ["維度", "KNN 系列", "Bridge 系列", "KDE 系列", "說明"]
for col, h in enumerate(headers3, 1):
    header_style(ws3.cell(row=2, column=col), h, bg="1B4F72")
ws3.row_dimensions[2].height = 18

comparison = [
    ("核心思想",      "找最相似的訓練樣本投票",   "用 SS 視窗建立橋樑相似度",  "估計測試點的物種空間密度",     "KDE 從「誰最像我」→「我在哪個分布裡」"),
    ("訓練資料量",    "66 個檔案",               "127,896 SS 視窗（橋）",     "739 訓練視窗",                 "KDE/Bridge 利用更多資料"),
    ("最佳 EP-only", "0.920（interaction_knn）", "N/A",                       "0.937（kde_per_species）",     "KDE 在 EP-only 也超越 KNN"),
    ("最佳全管線",    "0.9412",                   "0.9444",                    "0.9721（kde_perwin）",          "KDE 大幅超越前兩者"),
    ("天花板",        "0.920 EP-only",            "0.9444 全管線",             "持續提升中",                   "KDE 尚未見天花板"),
    ("推理複雜度",    "O(n_test × n_train)",      "O(n_test × n_bridge)",      "O(n_test × n_train × d)",      "KDE 在 d=32 時仍快"),
    ("需要外部資料",  "否",                        "是（soundscape）",           "否",                           "KDE 純依賴標注訓練集"),
    ("關鍵突破",      "per_species_alpha",        "SED species weighting",      "per-window scoring",           "每個方向的最重要改進"),
]

stage_colors = [C_STAGE1, C_STAGE2, C_STAGE3]
for i, row in enumerate(comparison, start=3):
    ws3.row_dimensions[i].height = 28
    dim, knn, bridge, kde, note = row
    bg = C_GRAY if i % 2 == 0 else C_WHITE
    data_style(ws3.cell(row=i, column=1), dim, bold=True, bg=bg)
    data_style(ws3.cell(row=i, column=2), knn, bg=C_STAGE1, wrap=True)
    data_style(ws3.cell(row=i, column=3), bridge, bg=C_STAGE2, wrap=True)
    data_style(ws3.cell(row=i, column=4), kde, bg=C_STAGE3, bold=True, wrap=True)
    data_style(ws3.cell(row=i, column=5), note, bg=bg, wrap=True)

ws3.freeze_panes = "A3"

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 4: 提交計畫 (Submission Plan)
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("提交計畫")

ws4.column_dimensions['A'].width = 6
ws4.column_dimensions['B'].width = 50
ws4.column_dimensions['C'].width = 14
ws4.column_dimensions['D'].width = 12
ws4.column_dimensions['E'].width = 38

ws4.merge_cells('A1:E1')
t4 = ws4['A1']
t4.value = "當前值得提交的 Notebooks（按優先順序）"
t4.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
t4.fill = PatternFill("solid", fgColor="1F3864")
t4.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 28

headers4 = ["優先", "Notebook 名稱", "CV AUC", "PKL 方法", "說明"]
for col, h in enumerate(headers4, 1):
    header_style(ws4.cell(row=2, column=col), h)
ws4.row_dimensions[2].height = 18

submissions = [
    (1, "dual-foundation-protossm-v6-kde-perwin.ipynb",  0.9721, "kde_perwin",          "每視窗 KDE → 平均；a=0.90, b=2.0 ← 目前最佳"),
    (2, "dual-foundation-protossm-v6-kde-rknn.ipynb",    0.9711, "kde_win_rknn_blend",  "視窗 KDE + RKNN k5；a=0.92, b=1.4"),
    (3, "dual-foundation-protossm-v6-kde-win.ipynb",     0.9701, "kde_window_level",    "視窗平均後 KDE + win_k1；a=0.90, b=1.2"),
    (4, "dual-foundation-protossm-v6-kde.ipynb",         0.9560, "kde_per_species",     "檔案級 KDE + win_k1；a=0.95, b=1.2"),
    (5, "v14-ls2-a090-b155（logspace geo+win）",         0.9408, "logspace_geo5_win1",  "舊版 LS2 系列；備用提交"),
    (6, "v14-win070-lam35（KNN window mix）",             0.9399, "attn_k4_win",         "舊版窗口 KNN 混合；最低優先"),
]

priority_colors = ["FF0000", "E67E22", "2E86C1", "1A7D44", "7D3C98", "808080"]
for i, (pri, nb, auc, pkl, note) in enumerate(submissions, start=3):
    ws4.row_dimensions[i].height = 22
    is_top = pri <= 3
    bg = "FFF3E0" if pri == 1 else (C_WHITE if pri <= 3 else C_GRAY)

    data_style(ws4.cell(row=i, column=1), pri, align="center", bold=is_top,
               color=priority_colors[pri-1], bg=bg)
    data_style(ws4.cell(row=i, column=2), nb, bold=is_top,
               color="C0392B" if pri == 1 else "000000", bg=bg)
    num_style(ws4.cell(row=i, column=3), auc, bold=is_top,
              color="C0392B" if pri == 1 else "000000", bg=bg)
    data_style(ws4.cell(row=i, column=4), pkl, bg=bg, align="center")
    data_style(ws4.cell(row=i, column=5), note, bg=bg, wrap=True)

ws4.freeze_panes = "A3"

# ── Save ───────────────────────────────────────────────────────────────────────
out_path = "reports/embed_prior_improvement_history.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
import os
size_kb = os.path.getsize(out_path) / 1024
print(f"Size: {size_kb:.1f} KB")
print(f"Sheets: {[s.title for s in wb.worksheets]}")
